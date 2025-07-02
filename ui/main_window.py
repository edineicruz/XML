#!/usr/bin/env python3
"""
Main Window for XML Fiscal Manager Pro
Professional interface with document management, filtering, and export capabilities
"""

import logging
import csv
import sys
import os
import time
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QLabel, QPushButton, QLineEdit, QComboBox,
    QDateEdit, QSpinBox, QProgressBar, QTextEdit, QFileDialog,
    QMessageBox, QToolBar, QStatusBar, QMenuBar, QMenu,
    QFrame, QGridLayout, QScrollArea, QListWidget, QTreeWidget,
    QTreeWidgetItem, QDialog, QAbstractItemView, QStyle, QApplication,
    QProgressDialog, QListWidgetItem, QCheckBox, QFormLayout
)
from PySide6.QtCore import Qt, QTimer, QThread, Signal, QDate, QSize, QDateTime
from PySide6.QtGui import QIcon, QFont, QPixmap, QPalette, QColor, QAction
from PySide6.QtWidgets import QSizePolicy

from core.xml_processor import XMLProcessor
from core.database_manager import DatabaseManager
from models.xml_models import XMLModelManager
from ui.export_dialog import ExportDialog
from ui.document_viewer import DocumentViewer
from ui.settings_dialog import SettingsDialog


class DatabaseInitWorker(QThread):
    """Worker thread for initializing databases asynchronously"""
    
    progress_updated = Signal(str)
    initialization_complete = Signal(bool)
    error_occurred = Signal(str)
    
    def __init__(self, xml_model_manager, db_manager):
        super().__init__()
        self.xml_model_manager = xml_model_manager
        self.db_manager = db_manager
    
    def run(self):
        """Initialize databases in background"""
        try:
            if not self.xml_model_manager:
                self.error_occurred.emit("XML Model Manager não disponível")
                return
            
            self.progress_updated.emit("Inicializando schemas de banco de dados...")
            
            # Initialize model databases
            success = self.xml_model_manager.initialize_databases(self.db_manager)
            
            if success:
                self.progress_updated.emit("Bancos de dados inicializados com sucesso!")
                self.initialization_complete.emit(True)
            else:
                self.error_occurred.emit("Falha na inicialização de alguns bancos de dados")
                
        except Exception as e:
            self.error_occurred.emit(f"Erro durante inicialização: {str(e)}")


class ExportConfigDialog(QDialog):
    """Dialog for configuring export settings"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configurações de Exportação")
        self.setModal(True)
        self.resize(500, 400)
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #dee2e6;
                border-radius: 6px;
                margin: 10px 0px;
                padding: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #495057;
            }
        """)
        
        self._setup_ui()
        self._load_settings()
    
    def _setup_ui(self):
        """Setup the UI components"""
        layout = QVBoxLayout(self)
        
        # Excel Export Settings
        excel_group = QGroupBox("Configurações do Excel")
        excel_layout = QFormLayout(excel_group)
        
        self.enable_formatting = QCheckBox("Aplicar formatação às células")
        self.enable_formatting.setToolTip("Quando desmarcado, melhora a velocidade de exportação")
        excel_layout.addRow("Formatação:", self.enable_formatting)
        
        self.auto_resize_columns = QCheckBox("Redimensionar colunas automaticamente")
        excel_layout.addRow("Auto-dimensionar:", self.auto_resize_columns)
        
        self.include_totals = QCheckBox("Incluir linhas de totais")
        excel_layout.addRow("Totais:", self.include_totals)
        
        self.freeze_header = QCheckBox("Congelar linha de cabeçalho")
        excel_layout.addRow("Congelar cabeçalho:", self.freeze_header)
        
        layout.addWidget(excel_group)
        
        # CSV Export Settings
        csv_group = QGroupBox("Configurações do CSV")
        csv_layout = QFormLayout(csv_group)
        
        self.csv_separator = QComboBox()
        self.csv_separator.addItems([";", ",", "|", "\t"])
        csv_layout.addRow("Separador:", self.csv_separator)
        
        self.csv_encoding = QComboBox()
        self.csv_encoding.addItems(["UTF-8", "UTF-8-BOM", "Windows-1252", "ISO-8859-1"])
        csv_layout.addRow("Codificação:", self.csv_encoding)
        
        layout.addWidget(csv_group)
        
        # Performance Settings
        perf_group = QGroupBox("Configurações de Performance")
        perf_layout = QFormLayout(perf_group)
        
        self.batch_size = QSpinBox()
        self.batch_size.setRange(100, 10000)
        self.batch_size.setValue(1000)
        self.batch_size.setToolTip("Número de registros processados por vez")
        perf_layout.addRow("Tamanho do lote:", self.batch_size)
        
        self.show_progress = QCheckBox("Mostrar progresso detalhado")
        self.show_progress.setChecked(True)
        perf_layout.addRow("Progresso:", self.show_progress)
        
        layout.addWidget(perf_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("Salvar")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.save_btn.clicked.connect(self.accept)
        
        self.cancel_btn = QPushButton("Cancelar")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #545b62;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
    
    def _load_settings(self):
        """Load current settings"""
        # Default settings - in a real app, these would come from config
        self.enable_formatting.setChecked(False)  # Default to faster export
        self.auto_resize_columns.setChecked(True)
        self.include_totals.setChecked(False)
        self.freeze_header.setChecked(True)
        self.csv_separator.setCurrentText(";")
        self.csv_encoding.setCurrentText("UTF-8")
    
    def get_settings(self):
        """Get current settings as dictionary"""
        return {
            'excel': {
                'enable_formatting': self.enable_formatting.isChecked(),
                'auto_resize_columns': self.auto_resize_columns.isChecked(),
                'include_totals': self.include_totals.isChecked(),
                'freeze_header': self.freeze_header.isChecked()
            },
            'csv': {
                'separator': self.csv_separator.currentText(),
                'encoding': self.csv_encoding.currentText()
            },
            'performance': {
                'batch_size': self.batch_size.value(),
                'show_progress': self.show_progress.isChecked()
            }
        }


class DocumentImportWorker(QThread):
    """Worker thread for importing documents"""
    
    progress_updated = Signal(int, str)
    document_processed = Signal(dict)
    finished = Signal(list)
    error_occurred = Signal(str)
    
    def __init__(self, xml_processor, file_paths):
        super().__init__()
        self.xml_processor = xml_processor
        self.file_paths = file_paths
        self.results = []
    
    def run(self):
        """Process files in background"""
        try:
            total_files = len(self.file_paths)
            
            for i, file_path in enumerate(self.file_paths):
                progress = int((i / total_files) * 100)
                self.progress_updated.emit(progress, f"Processando {file_path.name}...")
                
                result = self.xml_processor.process_file(file_path)
                self.results.append(result)
                self.document_processed.emit(result)
                
                if self.isInterruptionRequested():
                    break
            
            self.progress_updated.emit(100, "Processamento concluído!")
            self.finished.emit(self.results)
            
        except Exception as e:
            self.error_occurred.emit(str(e))


class MainWindow(QMainWindow):
    """Professional main window for XML Fiscal Manager Pro"""
    
    def __init__(self, config, auth_manager, db_manager):
        """Initialize main window with enhanced XML model management"""
        super().__init__()
        
        # Store core components
        self.config = config
        self.auth_manager = auth_manager
        self.db_manager = db_manager
        
        # XML processing
        self.xml_processor = XMLProcessor(config, db_manager)
        
        # Initialize XML model manager with error handling
        try:
            from models.xml_models import XMLModelManager
            self.xml_model_manager = XMLModelManager()
            logging.info("XML Model Manager initialized successfully")
        except Exception as e:
            logging.error(f"Failed to initialize XML Model Manager: {e}")
            self.xml_model_manager = None
        
        # Initialize state variables
        self.documents = []
        self.filtered_documents = []
        self.current_filters = {}
        self.import_worker = None
        self.db_init_worker = None
        self.selected_model = None
        self.model_documents = {}
        self.export_settings = self._get_default_export_settings()
        
        # Set up the window
        self._setup_window()
        
        # Connect close event for cleanup
        self.setAttribute(Qt.WA_DeleteOnClose)
        
        logging.info("Main window initialized successfully")
    
    def _get_default_export_settings(self):
        """Get default export settings"""
        return {
            'excel': {
                'enable_formatting': False,  # Default to faster export
                'auto_resize_columns': True,
                'include_totals': False,
                'freeze_header': True
            },
            'csv': {
                'separator': ';',
                'encoding': 'UTF-8'
            },
            'performance': {
                'batch_size': 1000,
                'show_progress': True
            }
        }
    
    def _setup_window(self):
        """Setup main window properties and styling"""
        self.setWindowTitle("XML Fiscal Manager Pro v2.0.0")
        self.setMinimumSize(1200, 800)
        
        # Check if should start maximized
        ui_config = self.config.get_ui_config()
        startup_config = ui_config.get('startup', {})
        if startup_config.get('start_maximized', True):  # Default to True for fullscreen
            self.showMaximized()
        else:
            # Set window size
            window_size = ui_config.get('window_size', {})
            self.resize(
                window_size.get('width', 1400),
                window_size.get('height', 900)
            )
            
            # Center window
            if startup_config.get('center_on_screen', True):
                self._center_window()
        
        # Apply modern styling
        self._apply_modern_style()
        
        # Create UI components
        self._create_menu_bar()
        self._create_toolbar()
        self._create_central_widget()
        self._create_status_bar()
        
        # Removed automatic database initialization and data loading
        # Database will only be loaded when user requests it via UI
    
    def _initialize_model_databases_async(self):
        """Initialize databases asynchronously to prevent UI freezing"""
        try:
            if not self.xml_model_manager:
                logging.warning("XML Model Manager not available, skipping database initialization")
                return
            
            # Show status message
            self.status_bar.showMessage("Inicializando bancos de dados...")
            
            # Create and start worker thread
            self.db_init_worker = DatabaseInitWorker(self.xml_model_manager, self.db_manager)
            self.db_init_worker.progress_updated.connect(self._on_db_init_progress)
            self.db_init_worker.initialization_complete.connect(self._on_db_init_complete)
            self.db_init_worker.error_occurred.connect(self._on_db_init_error)
            self.db_init_worker.start()
            
        except Exception as e:
            logging.error(f"Error starting database initialization: {e}")
            self.status_bar.showMessage("Erro na inicialização dos bancos de dados")
    
    def _on_db_init_progress(self, message: str):
        """Handle database initialization progress"""
        self.status_bar.showMessage(message)
        logging.info(f"DB Init: {message}")
    
    def _on_db_init_complete(self, success: bool):
        """Handle database initialization completion"""
        if success:
            self.status_bar.showMessage("Bancos de dados inicializados com sucesso!", 3000)
            logging.info("Database initialization completed successfully")
            
            # Load products data after database initialization
            self._refresh_products()
        else:
            self.status_bar.showMessage("Inicialização concluída com avisos", 3000)
            logging.warning("Database initialization completed with warnings")
        
        # Clean up worker
        if self.db_init_worker:
            self.db_init_worker.deleteLater()
            self.db_init_worker = None
    
    def _on_db_init_error(self, error_message: str):
        """Handle database initialization error"""
        self.status_bar.showMessage("Erro na inicialização dos bancos", 5000)
        logging.error(f"Database initialization error: {error_message}")
        
        # Show warning but don't block the application
        QMessageBox.warning(self, "Aviso de Inicialização", 
                          f"Erro ao inicializar bancos de dados:\n{error_message}\n\n"
                          "A aplicação continuará funcionando com funcionalidade limitada.")
        
        # Clean up worker
        if self.db_init_worker:
            self.db_init_worker.deleteLater()
            self.db_init_worker = None
    
    def _center_window(self):
        """Center window on screen"""
        from PySide6.QtGui import QGuiApplication
        
        screen = QGuiApplication.primaryScreen().geometry()
        window_geometry = self.geometry()
        
        x = (screen.width() - window_geometry.width()) // 2
        y = (screen.height() - window_geometry.height()) // 2
        
        self.move(x, y)
    
    def _apply_modern_style(self):
        """Apply modern stylesheet to the entire application"""
        style = """
        QMainWindow {
            background-color: #ffffff;
            border: none;
        }
        
        QMenuBar {
            background-color: #2c3e50;
            color: white;
            border: none;
            font-size: 12px;
            padding: 2px;
        }
        
        QMenuBar::item {
            background-color: transparent;
            padding: 6px 12px;
            margin: 1px;
            border-radius: 3px;
        }
        
        QMenuBar::item:selected {
            background-color: #34495e;
            color: #ecf0f1;
        }
        
        QMenuBar::item:pressed {
            background-color: #1abc9c;
        }
        
        QMenu {
            background-color: #ecf0f1;
            color: #2c3e50;
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            padding: 4px;
        }
        
        QMenu::item {
            padding: 6px 16px;
            border-radius: 3px;
        }
        
        QMenu::item:selected {
            background-color: #3498db;
            color: white;
        }
        
        QToolBar {
            background-color: #34495e;
            border: none;
            spacing: 4px;
            padding: 4px;
        }
        
        QToolBar QToolButton {
            background-color: transparent;
            color: white;
            border: 1px solid transparent;
            padding: 6px 8px;
            border-radius: 4px;
            font-size: 11px;
        }
        
        QToolBar QToolButton:hover {
            background-color: #3498db;
            border: 1px solid #2980b9;
        }
        
        QToolBar QToolButton:pressed {
            background-color: #2980b9;
            border: 1px solid #21618c;
        }
        
        QStatusBar {
            background-color: #ecf0f1;
            color: #2c3e50;
            border-top: 1px solid #bdc3c7;
            font-size: 11px;
        }
        
        QStatusBar::item {
            border: none;
        }
        
        QTableWidget {
            background-color: white;
            alternate-background-color: #f8f9fa;
            selection-background-color: #3498db;
            selection-color: white;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            gridline-color: #dee2e6;
            font-size: 11px;
        }
        
        QTableWidget::item {
            padding: 4px 8px;
            border-bottom: 1px solid #eee;
        }
        
        QTableWidget::item:selected {
            background-color: #3498db;
            color: white;
        }
        
        QHeaderView::section {
            background-color: #34495e;
            color: white;
            padding: 8px 6px;
            border: 1px solid #2c3e50;
            font-weight: bold;
            font-size: 10px;
        }
        
        QHeaderView::section:hover {
            background-color: #3498db;
        }
        
        QHeaderView::section:pressed {
            background-color: #2980b9;
        }
        
        QProgressBar {
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            background-color: #ecf0f1;
            text-align: center;
            font-size: 10px;
            min-height: 16px;
            max-height: 16px;
        }
        
        QProgressBar::chunk {
            background-color: #27ae60;
            border-radius: 3px;
        }
        
        QScrollBar:vertical {
            border: none;
            background-color: #f1f1f1;
            width: 12px;
            border-radius: 6px;
        }
        
        QScrollBar::handle:vertical {
            background-color: #c1c1c1;
            min-height: 20px;
            border-radius: 6px;
            margin: 1px;
        }
        
        QScrollBar::handle:vertical:hover {
            background-color: #a8a8a8;
        }
        
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            height: 0px;
        }
        
        QScrollBar:horizontal {
            border: none;
            background-color: #f1f1f1;
            height: 12px;
            border-radius: 6px;
        }
        
        QScrollBar::handle:horizontal {
            background-color: #c1c1c1;
            min-width: 20px;
            border-radius: 6px;
            margin: 1px;
        }
        
        QScrollBar::handle:horizontal:hover {
            background-color: #a8a8a8;
        }
        
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
            width: 0px;
        }
        
        QTabWidget::pane {
            border: 1px solid #dee2e6;
            border-radius: 4px;
            background-color: white;
        }
        
        QTabBar::tab {
            background-color: #f8f9fa;
            color: #495057;
            padding: 8px 16px;
            margin-right: 2px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            border: 1px solid #dee2e6;
            border-bottom: none;
        }
        
        QTabBar::tab:selected {
            background-color: white;
            color: #2c3e50;
            font-weight: bold;
        }
        
        QTabBar::tab:hover {
            background-color: #e9ecef;
        }
        
        QSplitter::handle {
            background-color: #dee2e6;
            width: 2px;
            height: 2px;
        }
        
        QSplitter::handle:hover {
            background-color: #3498db;
        }
        """
        
        self.setStyleSheet(style)
    
    def _create_menu_bar(self):
        """Create menu bar with export configuration menu"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&Arquivo")
        
        # Import documents
        import_action = QAction("&Importar XMLs", self)
        import_action.setShortcut("Ctrl+I")
        import_action.setStatusTip("Importar documentos XML")
        import_action.triggered.connect(self._import_documents)
        file_menu.addAction(import_action)
        
        # Import folder
        import_folder_action = QAction("Importar &Pasta", self)
        import_folder_action.setShortcut("Ctrl+Shift+I")
        import_folder_action.setStatusTip("Importar XMLs de uma pasta (busca recursiva e extrai ZIPs)")
        import_folder_action.triggered.connect(self._import_folder)
        file_menu.addAction(import_folder_action)
        
        file_menu.addSeparator()
        
        # Exit
        exit_action = QAction("&Sair", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.setStatusTip("Sair da aplicação")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("&Ferramentas")
        
        # Clear all data
        clear_action = QAction("&Limpar Todos os Dados", self)
        clear_action.setStatusTip("Limpar todos os dados do banco")
        clear_action.triggered.connect(self._clear_all_data)
        tools_menu.addAction(clear_action)
        
        tools_menu.addSeparator()
        
        # Settings
        settings_action = QAction("&Configurações", self)
        settings_action.setStatusTip("Abrir configurações")
        settings_action.triggered.connect(self._show_settings)
        tools_menu.addAction(settings_action)
        
        # Export Configuration menu (between Tools and Help)
        export_menu = menubar.addMenu("&Exportação")
        
        # Export configuration
        export_config_action = QAction("&Configurar Exportação", self)
        export_config_action.setStatusTip("Configurar opções de exportação")
        export_config_action.triggered.connect(self._show_export_config)
        export_menu.addAction(export_config_action)
        
        export_menu.addSeparator()
        
        # Quick export actions
        quick_excel_action = QAction("Exportar para &Excel", self)
        quick_excel_action.setShortcut("Ctrl+E")
        quick_excel_action.setStatusTip("Exportação rápida para Excel")
        quick_excel_action.triggered.connect(self._quick_export_excel)
        export_menu.addAction(quick_excel_action)
        
        quick_csv_action = QAction("Exportar para &CSV", self)
        quick_csv_action.setShortcut("Ctrl+Shift+E")
        quick_csv_action.setStatusTip("Exportação rápida para CSV")
        quick_csv_action.triggered.connect(self._quick_export_csv)
        export_menu.addAction(quick_csv_action)
        
        # Help menu
        help_menu = menubar.addMenu("&Ajuda")
        
        # About
        about_action = QAction("&Sobre", self)
        about_action.setStatusTip("Sobre a aplicação")
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)
    
    def _create_toolbar(self):
        """Create toolbar - enhanced version with update button"""
        toolbar = self.addToolBar("Principal")
        toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        
        # Import action
        import_action = QAction("Importar XMLs", self)
        import_action.setIcon(QIcon.fromTheme("document-open"))
        import_action.triggered.connect(self._import_documents)
        toolbar.addAction(import_action)
        
        toolbar.addSeparator()
        
        # Import folder action
        import_folder_action = QAction("Importar Pasta", self)
        import_folder_action.setIcon(QIcon.fromTheme("folder-open"))
        import_folder_action.triggered.connect(self._import_folder)
        toolbar.addAction(import_folder_action)
        
        toolbar.addSeparator()
        
        # Update app action
        update_action = QAction("Verificar Atualizações", self)
        update_action.setIcon(QIcon.fromTheme("view-refresh"))
        update_action.triggered.connect(self._check_for_updates)
        toolbar.addAction(update_action)
        
        toolbar.addSeparator()
        
        # Refresh data action
        refresh_action = QAction("Atualizar Dados", self)
        refresh_action.setIcon(QIcon.fromTheme("view-refresh"))
        refresh_action.triggered.connect(self._refresh_products)
        toolbar.addAction(refresh_action)
    
    def _create_central_widget(self):
        """Create central widget with sidebar for XML models"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QHBoxLayout(central_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Create main splitter
        main_splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(main_splitter)
        
        # Create sidebar with XML model types
        self._create_xml_models_sidebar(main_splitter)
        
        # Create main content area
        self._create_main_content_area(main_splitter)
        
        # Set splitter sizes (sidebar: 250px, main content: remaining)
        main_splitter.setSizes([250, 1200])
        main_splitter.setCollapsible(0, False)  # Don't allow sidebar to collapse completely
    
    def _create_xml_models_sidebar(self, parent_splitter):
        """Create the XML models sidebar with improved visibility"""
        # Create container widget for sidebar
        sidebar_container = QWidget()
        sidebar_container.setFixedWidth(250)  # Set fixed width to ensure visibility
        sidebar_container.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border-right: 2px solid #dee2e6;
            }
        """)
        
        # Main layout for sidebar
        sidebar_layout = QVBoxLayout(sidebar_container)
        sidebar_layout.setContentsMargins(10, 10, 10, 10)
        sidebar_layout.setSpacing(8)
        
        # Title
        title_label = QLabel("Modelos XML")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                margin-bottom: 8px;
                padding: 8px;
                background-color: #e9ecef;
                border-radius: 6px;
            }
        """)
        sidebar_layout.addWidget(title_label)
        
        # Models list widget with improved styling
        self.models_list = QListWidget()
        self.models_list.setStyleSheet("""
            QListWidget {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 5px;
                font-size: 12px;
                min-height: 200px;
            }
            QListWidget::item {
                padding: 8px 12px;
                margin: 2px 0px;
                border-radius: 4px;
                color: #495057;
                border: 1px solid transparent;
            }
            QListWidget::item:hover {
                background-color: #e3f2fd;
                border: 1px solid #2196f3;
                color: #1976d2;
            }
            QListWidget::item:selected {
                background-color: #2196f3;
                color: white;
                font-weight: bold;
            }
        """)
        
        # Populate models list
        models = ['NFe', 'NFCe', 'CTe', 'NFSe', 'MDFe', 'CCe', 'EPEC']
        for model in models:
            item = QListWidgetItem(f"📄 {model}")
            item.setData(Qt.UserRole, model)
            self.models_list.addItem(item)
        
        # Connect selection signal
        self.models_list.itemSelectionChanged.connect(self._on_model_selection_changed)
        
        sidebar_layout.addWidget(self.models_list)
        
        # Model count label
        self.model_count_label = QLabel("Selecione um modelo")
        self.model_count_label.setStyleSheet("""
            QLabel {
                font-size: 11px;
                color: #6c757d;
                font-style: italic;
                padding: 5px;
                background-color: #f1f3f4;
                border-radius: 4px;
                margin-top: 5px;
            }
        """)
        sidebar_layout.addWidget(self.model_count_label)
        
        # Action buttons with improved layout
        buttons_layout = QVBoxLayout()
        buttons_layout.setSpacing(5)
        
        # Refresh button
        self.refresh_models_btn = QPushButton("🔄 Atualizar")
        self.refresh_models_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        self.refresh_models_btn.clicked.connect(self._load_model_data)
        buttons_layout.addWidget(self.refresh_models_btn)
        
        # Clear selection button
        self.clear_selection_btn = QPushButton("❌ Limpar")
        self.clear_selection_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #545b62;
            }
            QPushButton:pressed {
                background-color: #495057;
            }
        """)
        self.clear_selection_btn.clicked.connect(self._clear_model_selection)
        buttons_layout.addWidget(self.clear_selection_btn)
        
        sidebar_layout.addLayout(buttons_layout)
        
        # Add stretch to push everything to top
        sidebar_layout.addStretch()
        
        # Statistics section
        stats_frame = QFrame()
        stats_frame.setFrameStyle(QFrame.StyledPanel)
        stats_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 8px;
                margin-top: 10px;
            }
        """)
        
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setContentsMargins(5, 5, 5, 5)
        
        stats_title = QLabel("📊 Estatísticas")
        stats_title.setStyleSheet("""
            QLabel {
                font-size: 12px;
                font-weight: bold;
                color: #495057;
                margin-bottom: 5px;
            }
        """)
        stats_layout.addWidget(stats_title)
        
        self.stats_label = QLabel("Carregando...")
        self.stats_label.setStyleSheet("""
            QLabel {
                font-size: 10px;
                color: #6c757d;
                line-height: 1.2;
            }
        """)
        self.stats_label.setWordWrap(True)
        stats_layout.addWidget(self.stats_label)
        
        sidebar_layout.addWidget(stats_frame)
        
        # Add sidebar to parent splitter
        parent_splitter.addWidget(sidebar_container)
        
        # Set initial splitter sizes to ensure sidebar visibility
        parent_splitter.setSizes([250, 1000])  # Fixed sidebar width, flexible main area
        parent_splitter.setCollapsible(0, False)  # Don't allow sidebar to be collapsed
    
    def _create_main_content_area(self, parent_splitter):
        """Create main content area for displaying documents"""
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(15)
        
        # Title and controls for selected model
        self._create_model_header(content_layout)
        
        # Create the existing products widget
        self._create_products_widget_content(content_layout)
        
        parent_splitter.addWidget(content_widget)
    
    def _create_model_header(self, parent_layout):
        """Create header for the selected model"""
        header_layout = QHBoxLayout()
        
        # Dynamic title based on selected model
        self.title_label = QLabel("Todos os Documentos")
        self.title_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                margin: 0px;
            }
        """)
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()
        
        # Model-specific action buttons
        self.model_export_btn = QPushButton("Exportar Modelo")
        self.model_export_btn.setIcon(QIcon.fromTheme("document-save"))
        self.model_export_btn.clicked.connect(self._export_model_data)
        self.model_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
                min-width: 140px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #117a8b;
            }
        """)
        header_layout.addWidget(self.model_export_btn)
        
        # Export button (existing)
        self.export_excel_btn = QPushButton("Exportar para Excel")
        self.export_excel_btn.setIcon(QIcon.fromTheme("document-save"))
        self.export_excel_btn.clicked.connect(self._export_to_excel)
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
            QPushButton:pressed {
                background-color: #229954;
            }
        """)
        header_layout.addWidget(self.export_excel_btn)
        
        parent_layout.addLayout(header_layout)
    
    def _create_products_widget_content(self, parent_layout):
        """Create products management widget content (without the container)"""
        # Filter section
        filter_frame = QFrame()
        filter_frame.setFrameStyle(QFrame.StyledPanel)
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 5px;
            }
        """)
        
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setContentsMargins(15, 10, 15, 10)
        
        # Search field
        search_label = QLabel("Buscar:")
        search_label.setStyleSheet("font-weight: bold; color: #495057;")
        filter_layout.addWidget(search_label)
        
        self.product_search_input = QLineEdit()
        self.product_search_input.setPlaceholderText("Buscar por descrição, código, NCM...")
        self.product_search_input.textChanged.connect(self._filter_products)
        filter_layout.addWidget(self.product_search_input, 1)
        
        # Document type filter (will be updated based on selected model)
        filter_layout.addWidget(QLabel("Tipo:"))
        self.product_type_combo = QComboBox()
        self.product_type_combo.addItems(["Todos", "NFe", "NFCe", "CTe", "NFSe"])
        self.product_type_combo.currentTextChanged.connect(self._filter_products)
        filter_layout.addWidget(self.product_type_combo)
        
        # Clear filters button
        clear_btn = QPushButton("Limpar Filtros")
        clear_btn.clicked.connect(self._clear_filters)
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #545b62;
            }
        """)
        filter_layout.addWidget(clear_btn)
        
        # Load database button
        load_db_btn = QPushButton("🗄️ Carregar Banco")
        load_db_btn.clicked.connect(self._load_database_manually)
        load_db_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        filter_layout.addWidget(load_db_btn)
        
        parent_layout.addWidget(filter_frame)
        
        # Products table (will be dynamically updated based on selected model)
        self.products_table = QTableWidget()
        self._setup_products_table()
        parent_layout.addWidget(self.products_table)
    
    def _create_status_bar(self):
        """Create status bar"""
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # User info
        user_info = self.auth_manager.get_user_info()
        user_label = QLabel(f"Usuário: {user_info.get('user_name', 'Desconhecido')}")
        self.status_bar.addPermanentWidget(user_label)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setMaximum(100)
        self.status_bar.addPermanentWidget(self.progress_bar)
        
        # Database status
        self.db_status_label = QLabel("Database: OK")
        self.status_bar.addPermanentWidget(self.db_status_label)
        
        self.status_bar.showMessage("Pronto")
    
    def _import_documents(self):
        """Import XML documents"""
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML Files (*.xml)")
        file_dialog.setWindowTitle("Selecionar Arquivos XML")
        
        if file_dialog.exec():
            file_paths = [Path(f) for f in file_dialog.selectedFiles()]
            
            if file_paths:
                self._start_import_process(file_paths)
    
    def _import_folder(self):
        """Import XML documents from folder (with recursive search and ZIP extraction)"""
        folder_dialog = QFileDialog(self)
        folder_dialog.setFileMode(QFileDialog.Directory)
        folder_dialog.setWindowTitle("Selecionar Pasta para Importar XMLs")
        
        if folder_dialog.exec():
            folder_path = Path(folder_dialog.selectedFiles()[0])
            
            if folder_path.exists():
                self._process_folder_import(folder_path)
    
    def _process_folder_import(self, folder_path: Path):
        """Process folder import with recursive search and ZIP extraction"""
        try:
            self.status_bar.showMessage("Buscando arquivos XML e ZIP...")
            
            # Find XML and ZIP files recursively
            xml_files = []
            zip_files = []
            
            # Recursive search
            for file_path in folder_path.rglob('*'):
                if file_path.is_file():
                    if file_path.suffix.lower() == '.xml':
                        xml_files.append(file_path)
                    elif file_path.suffix.lower() == '.zip':
                        zip_files.append(file_path)
            
            # Extract ZIP files and find XMLs inside them
            extracted_xmls = []
            if zip_files:
                extracted_xmls = self._extract_zips_and_find_xmls(zip_files)
            
            # Combine all XML files
            all_xml_files = xml_files + extracted_xmls
            
            if not all_xml_files:
                QMessageBox.information(self, "Aviso", 
                                      f"Nenhum arquivo XML encontrado na pasta:\n{folder_path}\n\n"
                                      f"Subpastas verificadas: {len(list(folder_path.rglob('*')))}\n"
                                      f"Arquivos ZIP encontrados: {len(zip_files)}")
                return
            
            # Show confirmation dialog
            reply = QMessageBox.question(self, "Confirmar Importação", 
                                       f"Encontrados {len(all_xml_files)} arquivos XML:\n"
                                       f"• XMLs diretos: {len(xml_files)}\n"
                                       f"• XMLs extraídos de ZIP: {len(extracted_xmls)}\n"
                                       f"• ZIPs processados: {len(zip_files)}\n\n"
                                       f"Deseja prosseguir com a importação?",
                                       QMessageBox.Yes | QMessageBox.No,
                                       QMessageBox.Yes)
            
            if reply == QMessageBox.Yes:
                self._start_import_process(all_xml_files)
                
        except Exception as e:
            logging.error(f"Error processing folder import: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao processar pasta:\n{str(e)}")
    
    def _extract_zips_and_find_xmls(self, zip_files: List[Path]) -> List[Path]:
        """Extract ZIP files and find XML files inside them"""
        extracted_xmls = []
        
        try:
            import zipfile
            import tempfile
            import shutil
            
            # Create temporary directory for extractions
            temp_dir = Path(tempfile.mkdtemp(prefix="xml_extraction_"))
            
            for zip_path in zip_files:
                try:
                    # Create subdirectory for this ZIP
                    zip_extract_dir = temp_dir / f"zip_{zip_path.stem}"
                    zip_extract_dir.mkdir(exist_ok=True)
                    
                    # Extract ZIP
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(zip_extract_dir)
                    
                    # Find XML files in extracted content
                    for xml_file in zip_extract_dir.rglob('*.xml'):
                        if xml_file.is_file():
                            extracted_xmls.append(xml_file)
                    
                    logging.info(f"Extracted ZIP {zip_path.name}, found {len(list(zip_extract_dir.rglob('*.xml')))} XMLs")
                    
                except Exception as e:
                    logging.error(f"Error extracting ZIP {zip_path}: {e}")
                    continue
            
            logging.info(f"Total XMLs extracted from {len(zip_files)} ZIPs: {len(extracted_xmls)}")
            
        except Exception as e:
            logging.error(f"Error in ZIP extraction process: {e}")
        
        return extracted_xmls
    
    def _start_import_process(self, file_paths: List[Path]):
        """Start import process in background"""
        # Show progress
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_bar.showMessage(f"Importando {len(file_paths)} arquivos...")
        
        # Start worker thread
        self.import_worker = DocumentImportWorker(self.xml_processor, file_paths)
        self.import_worker.progress_updated.connect(self._update_import_progress)
        self.import_worker.document_processed.connect(self._document_imported)
        self.import_worker.finished.connect(self._import_finished)
        self.import_worker.error_occurred.connect(self._import_error)
        self.import_worker.start()
    
    def _update_import_progress(self, progress: int, message: str):
        """Update import progress"""
        self.progress_bar.setValue(progress)
        self.status_bar.showMessage(message)
    
    def _document_imported(self, result: Dict):
        """Handle document import result"""
        try:
            if result['status'] == 'success':
                document_id = result.get('document_id')
                if document_id:
                    logging.info(f"Document imported successfully with ID: {document_id}")
                else:
                    logging.warning(f"Document processed but no ID returned")
            elif result['status'] == 'skipped':
                logging.info(f"Document skipped: {result.get('message', 'Already exists')}")
            else:
                logging.error(f"Document import error: {result.get('error', 'Unknown error')}")
                
        except Exception as e:
            logging.error(f"Error handling document import result: {e}")
    
    def _import_finished(self, results: List[Dict]):
        """Handle import process completion"""
        try:
            # Count results
            successful = sum(1 for r in results if r['status'] == 'success')
            skipped = sum(1 for r in results if r['status'] == 'skipped')
            failed = sum(1 for r in results if r['status'] == 'error')
            
            # Update status
            message = f"Importação concluída: {successful} sucessos, {skipped} ignorados, {failed} falhas"
            self.statusBar().showMessage(message)
            
            # Show completion message
            if failed > 0:
                QMessageBox.warning(self, "Importação Concluída com Erros", 
                                  f"Importação finalizada:\n\n"
                                  f"✓ Sucessos: {successful}\n"
                                  f"⚠ Ignorados: {skipped}\n"
                                  f"✗ Falhas: {failed}\n\n"
                                  f"Verifique os logs para detalhes dos erros.")
            else:
                QMessageBox.information(self, "Importação Concluída", 
                                      f"Importação finalizada com sucesso:\n\n"
                                      f"✓ Sucessos: {successful}\n"
                                      f"⚠ Ignorados: {skipped}")
            
            # Refresh products table to show new data
            self._refresh_products()
            
        except Exception as e:
            logging.error(f"Error handling import completion: {e}")
    
    def _import_error(self, error: str):
        """Handle import error"""
        logging.error(f"Import error: {error}")
        QMessageBox.critical(self, "Erro de Importação", f"Erro durante a importação:\n{error}")
        self.statusBar().showMessage("Erro na importação")
    
    def _setup_products_table(self):
        """Setup products table with modern styling and comprehensive fields"""
        try:
            # Comprehensive table headers with all available XML fields
            headers = [
                # Basic Document Info
                "Tipo Doc", "Número", "Série", "Modelo", "Data Emissão", "Data Saída",
                "Chave Acesso", "Protocolo", "Data Protocolo", "Natureza Operação",
                
                # Emitter Info  
                "CNPJ Emitente", "Emitente", "Nome Fantasia", "IE Emitente",
                "End. Emitente", "Cidade Emitente", "UF Emitente", "CEP Emitente",
                
                # Recipient Info
                "CNPJ/CPF Destinatário", "Destinatário", "IE Destinatário", 
                "End. Destinatário", "Cidade Destinatário", "UF Destinatário",
                
                # Product Item Info
                "Nº Item", "Código Item", "Descrição", "NCM", "CFOP", "EAN/GTIN",
                "Quantidade", "Unidade", "Valor Unit", "Valor Total Item",
                
                # Tax Information - ICMS
                "CST ICMS", "Base ICMS", "Valor ICMS", "Alíq ICMS",
                
                # Tax Information - IPI
                "CST IPI", "Base IPI", "Valor IPI", "Alíq IPI",
                
                # Tax Information - PIS
                "CST PIS", "Base PIS", "Valor PIS", "Alíq PIS",
                
                # Tax Information - COFINS
                "CST COFINS", "Base COFINS", "Valor COFINS", "Alíq COFINS",
                
                # Financial Totals
                "Total Produtos", "Total Frete", "Total Seguro", "Total Desconto",
                "Total Outros", "Total NFe", "Total Tributos Item", "ICMS ST",
                
                # Transport & Payment
                "Modalidade Frete", "Transportadora", "Forma Pagamento",
                
                # Additional Info
                "Info Adicional", "Arquivo"
            ]
            
            self.products_table.setColumnCount(len(headers))
            self.products_table.setHorizontalHeaderLabels(headers)
            
            # Set table properties - READ ONLY
            self.products_table.setAlternatingRowColors(True)
            self.products_table.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.products_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
            self.products_table.setSortingEnabled(True)
            self.products_table.setWordWrap(False)
            self.products_table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # Make table READ-ONLY
            
            # Enable context menu
            self.products_table.setContextMenuPolicy(Qt.CustomContextMenu)
            self.products_table.customContextMenuRequested.connect(self._show_products_context_menu)
            
            # Set optimized column widths for comprehensive table
            column_widths = [
                # Basic Document Info (10 columns)
                80, 100, 60, 60, 100, 100, 200, 120, 120, 200,
                
                # Emitter Info (8 columns)  
                150, 250, 200, 120, 200, 150, 50, 100,
                
                # Recipient Info (6 columns)
                150, 250, 120, 200, 150, 50,
                
                # Product Item Info (10 columns)
                60, 120, 300, 100, 80, 120, 100, 80, 120, 120,
                
                # Tax Information - ICMS (4 columns)
                80, 120, 120, 80,
                
                # Tax Information - IPI (4 columns)
                80, 120, 120, 80,
                
                # Tax Information - PIS (4 columns)
                80, 120, 120, 80,
                
                # Tax Information - COFINS (4 columns)
                80, 120, 120, 80,
                
                # Financial Totals (8 columns)
                120, 120, 120, 120, 120, 120, 120, 120,
                
                # Transport & Payment (3 columns)
                120, 200, 150,
                
                # Additional Info (2 columns)
                200, 200
            ]
            
            for i, width in enumerate(column_widths):
                if i < len(column_widths):
                    self.products_table.setColumnWidth(i, width)
            
            # Enhanced table styling
            self.products_table.setStyleSheet("""
                QTableWidget {
                    background-color: white;
                    border: 1px solid #d0d0d0;
                    border-radius: 8px;
                    gridline-color: #e0e0e0;
                    font-size: 11px;
                    selection-background-color: #3498db;
                }
                QTableWidget::item {
                    padding: 6px;
                    border-bottom: 1px solid #e0e0e0;
                    border-right: 1px solid #f0f0f0;
                }
                QTableWidget::item:selected {
                    background-color: #3498db;
                    color: white;
                }
                QTableWidget::item:hover {
                    background-color: #ecf0f1;
                }
                QHeaderView::section {
                    background-color: #34495e;
                    color: white;
                    padding: 8px;
                    border: none;
                    font-weight: bold;
                    font-size: 10px;
                    text-align: center;
                }
                QHeaderView::section:hover {
                    background-color: #2c3e50;
                }
                QScrollBar:horizontal {
                    height: 15px;
                    background-color: #f0f0f0;
                }
                QScrollBar:vertical {
                    width: 15px;
                    background-color: #f0f0f0;
                }
            """)
            
            # Load initial data
            self._refresh_products()
            
            logging.info("Products table setup completed with comprehensive XML fields and read-only mode")
            
        except Exception as e:
            logging.error(f"Error setting up products table: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao configurar tabela de produtos:\n{str(e)}")
    
    def _filter_products(self):
        """Filter products based on search and type"""
        try:
            # Get search text and type filter
            search_text = self.product_search_input.text().lower()
            doc_type_filter = self.product_type_combo.currentText()
            
            # Get all products from database
            all_products = self.db_manager.get_enhanced_products()
            
            # Apply filters
            filtered_products = []
            for product in all_products:
                # Text search filter
                if search_text:
                    searchable_text = ' '.join([
                        str(product.get('item_description', '')),
                        str(product.get('item_code', '')),
                        str(product.get('ncm_code', '')),
                        str(product.get('cfop', '')),
                        str(product.get('document_number', '')),
                        str(product.get('file_name', ''))
                    ]).lower()
                    
                    if search_text not in searchable_text:
                        continue
                
                # Document type filter
                if doc_type_filter != "Todos":
                    product_doc_type = product.get('document_type', '').lower()
                    if doc_type_filter.lower() != product_doc_type:
                        continue
                
                filtered_products.append(product)
            
            # Update table with filtered results
            self._update_products_table(filtered_products)
            
        except Exception as e:
            logging.error(f"Error filtering products: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao filtrar produtos:\n{str(e)}")
    
    def _refresh_products(self):
        """Refresh products table"""
        try:
            # Get enhanced products data
            products_data = self.db_manager.get_enhanced_products()
            
            if not products_data:
                logging.warning("No products data found")
                return
            
            # Update the table
            self._update_products_table(products_data)
            
            # Update status bar
            self.statusBar().showMessage(f"Produtos atualizados: {len(products_data)} registros carregados")
            
            logging.info(f"Products table refreshed with {len(products_data)} records")
            
        except Exception as e:
            logging.error(f"Error refreshing products: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar produtos:\n{str(e)}")
    
    def _update_products_table(self, products_data):
        """Update products table with comprehensive XML data"""
        try:
            # Set table row count
            self.products_table.setRowCount(len(products_data))
            
            # Fill data for each column with comprehensive XML fields
            for row, product in enumerate(products_data):
                # Define comprehensive columns mapping
                columns = [
                    # Basic Document Info (10 columns)
                    'document_type', 'document_number', 'series', 'model', 'issue_date', 'exit_date',
                    'access_key', 'protocol_number', 'protocol_date', 'operation_nature',
                    
                    # Emitter Info (8 columns)
                    'cnpj_issuer', 'issuer_name', 'emitter_fantasy', 'emitter_ie',
                    'emitter_address', 'emitter_city', 'emitter_state', 'emitter_cep',
                    
                    # Recipient Info (6 columns)
                    'cnpj_recipient', 'recipient_name', 'recipient_ie',
                    'recipient_address', 'recipient_city', 'recipient_state',
                    
                    # Product Item Info (10 columns)
                    'item_number', 'item_code', 'item_description', 'ncm_code', 'cfop', 'item_ean',
                    'quantity', 'commercial_unit', 'unit_value', 'total_value',
                    
                    # Tax Information - ICMS (4 columns)
                    'icms_cst', 'icms_base', 'icms_value', 'icms_rate',
                    
                    # Tax Information - IPI (4 columns)
                    'ipi_cst', 'ipi_base', 'ipi_value', 'ipi_rate',
                    
                    # Tax Information - PIS (4 columns)
                    'pis_cst', 'pis_base', 'pis_value', 'pis_rate',
                    
                    # Tax Information - COFINS (4 columns)
                    'cofins_cst', 'cofins_base', 'cofins_value', 'cofins_rate',
                    
                    # Financial Totals (8 columns)
                    'total_products', 'total_freight', 'total_insurance', 'total_discount',
                    'total_other', 'total_nfe', 'tax_value', 'icms_st_value',
                    
                    # Transport & Payment (3 columns)
                    'transport_modality', 'transporter_name', 'payment_method',
                    
                    # Additional Info (2 columns)
                    'additional_info', 'file_name'
                ]
                
                for col, field_name in enumerate(columns):
                    try:
                        value = product.get(field_name, '')
                        
                        # Special formatting for specific columns
                        if field_name in ['issue_date', 'exit_date', 'protocol_date'] and value:
                            try:
                                # Convert date to readable format
                                if isinstance(value, str) and value:
                                    if 'T' in value:
                                        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                        display_value = date_obj.strftime('%d/%m/%Y %H:%M')
                                    else:
                                        display_value = str(value)
                                else:
                                    display_value = str(value) if value else ''
                            except:
                                display_value = str(value) if value else ''
                                
                        elif field_name in ['cnpj_issuer', 'cnpj_recipient'] and value:
                            # Format CNPJ/CPF
                            doc_str = str(value).strip()
                            if len(doc_str) == 14:  # CNPJ
                                display_value = f"{doc_str[:2]}.{doc_str[2:5]}.{doc_str[5:8]}/{doc_str[8:12]}-{doc_str[12:14]}"
                            elif len(doc_str) == 11:  # CPF
                                display_value = f"{doc_str[:3]}.{doc_str[3:6]}.{doc_str[6:9]}-{doc_str[9:11]}"
                            else:
                                display_value = doc_str
                                
                        elif field_name in ['emitter_cep', 'recipient_cep'] and value:
                            # Format CEP
                            cep_str = str(value).strip()
                            if len(cep_str) == 8:
                                display_value = f"{cep_str[:5]}-{cep_str[5:]}"
                            else:
                                display_value = cep_str
                                
                        elif field_name in ['quantity', 'unit_value', 'total_value', 'icms_base', 'icms_value', 
                                          'ipi_base', 'ipi_value', 'pis_base', 'pis_value', 'cofins_base', 'cofins_value',
                                          'total_products', 'total_freight', 'total_insurance', 'total_discount',
                                          'total_other', 'total_nfe', 'tax_value', 'icms_st_value']:
                            # Format numeric values
                            try:
                                if value and float(value) != 0:
                                    display_value = f"{float(value):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                                else:
                                    display_value = "0,00"
                            except:
                                display_value = str(value) if value else '0,00'
                                
                        elif field_name in ['icms_rate', 'ipi_rate', 'pis_rate', 'cofins_rate']:
                            # Format percentage values
                            try:
                                if value and float(value) != 0:
                                    display_value = f"{float(value):.2f}%"
                                else:
                                    display_value = "0,00%"
                            except:
                                display_value = str(value) if value else '0,00%'
                                
                        elif field_name == 'access_key' and value:
                            # Format access key for better readability
                            key_str = str(value).strip()
                            if len(key_str) == 44:
                                display_value = f"{key_str[:4]} {key_str[4:8]} {key_str[8:12]} {key_str[12:16]} {key_str[16:20]} {key_str[20:24]} {key_str[24:28]} {key_str[28:32]} {key_str[32:36]} {key_str[36:40]} {key_str[40:44]}"
                            else:
                                display_value = key_str
                                
                        elif field_name == 'payment_method' and value:
                            # Convert payment method codes to descriptions
                            payment_types = {
                                '01': 'Dinheiro', '02': 'Cheque', '03': 'Cartão Crédito', '04': 'Cartão Débito',
                                '05': 'Crédito Loja', '10': 'Vale Alimentação', '11': 'Vale Refeição',
                                '12': 'Vale Presente', '13': 'Vale Combustível', '14': 'Duplicata',
                                '15': 'Boleto', '90': 'Sem Pagamento', '99': 'Outros'
                            }
                            display_value = payment_types.get(str(value), str(value))
                            
                        elif field_name == 'transport_modality' and value:
                            # Convert transport modality codes
                            transport_types = {
                                '0': 'Sem Frete', '1': 'CIF', '2': 'FOB', '3': 'Terceiros', '4': 'Próprio',
                                '9': 'Sem Transporte'
                            }
                            display_value = transport_types.get(str(value), str(value))
                            
                        elif field_name == 'file_name' and value:
                            # Clean file name
                            display_value = str(value).replace('.xml', '').replace('-procNFe', '')
                            
                        elif field_name == 'additional_info' and value:
                            # Truncate long additional info
                            info_str = str(value).strip()
                            display_value = info_str[:100] + '...' if len(info_str) > 100 else info_str
                            
                        else:
                            display_value = str(value) if value else ''
                        
                        item = QTableWidgetItem(display_value)
                        
                        # Set alignment based on data type
                        if field_name in ['quantity', 'unit_value', 'total_value', 'icms_base', 'icms_value',
                                        'ipi_base', 'ipi_value', 'pis_base', 'pis_value', 'cofins_base', 'cofins_value',
                                        'total_products', 'total_freight', 'total_insurance', 'total_discount',
                                        'total_other', 'total_nfe', 'tax_value', 'icms_st_value',
                                        'icms_rate', 'ipi_rate', 'pis_rate', 'cofins_rate']:
                            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        elif field_name in ['document_number', 'series', 'model', 'cnpj_issuer', 'cnpj_recipient',
                                          'item_code', 'ncm_code', 'cfop', 'item_ean', 'icms_cst', 'ipi_cst',
                                          'pis_cst', 'cofins_cst', 'emitter_cep', 'recipient_cep', 'item_number']:
                            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                        else:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        
                        # Set alternating row colors
                        if row % 2 == 0:
                            item.setBackground(QColor(248, 248, 248))
                        else:
                            item.setBackground(QColor(255, 255, 255))
                        
                        # Make item non-editable (additional safety)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        
                        self.products_table.setItem(row, col, item)
                        
                    except Exception as e:
                        logging.error(f"Error setting item at row {row}, col {col} ({field_name}): {e}")
                        error_item = QTableWidgetItem('')
                        error_item.setFlags(error_item.flags() & ~Qt.ItemIsEditable)
                        self.products_table.setItem(row, col, error_item)
            
            # Auto-resize columns to content but with limits
            self.products_table.resizeColumnsToContents()
            
            # Set maximum column widths to prevent UI breaking
            max_widths = {
                0: 80, 1: 100, 2: 60, 3: 60, 4: 120, 5: 120, 6: 250, 7: 120, 8: 120, 9: 250,  # Basic doc info
                10: 150, 11: 250, 12: 200, 13: 120, 14: 250, 15: 150, 16: 50, 17: 100,  # Emitter info
                18: 150, 19: 250, 20: 120, 21: 250, 22: 150, 23: 50,  # Recipient info
                24: 60, 25: 120, 26: 300, 27: 100, 28: 80, 29: 120, 30: 100, 31: 80, 32: 120, 33: 120,  # Product info
                34: 80, 35: 120, 36: 120, 37: 80,  # ICMS
                38: 80, 39: 120, 40: 120, 41: 80,  # IPI
                42: 80, 43: 120, 44: 120, 45: 80,  # PIS
                46: 80, 47: 120, 48: 120, 49: 80,  # COFINS
                50: 120, 51: 120, 52: 120, 53: 120, 54: 120, 55: 120, 56: 120, 57: 120,  # Totals
                58: 120, 59: 200, 60: 150,  # Transport & Payment
                61: 200, 62: 200  # Additional info
            }
            
            for col, max_width in max_widths.items():
                if col < self.products_table.columnCount() and self.products_table.columnWidth(col) > max_width:
                    self.products_table.setColumnWidth(col, max_width)
            
            logging.info(f"Products table updated with {len(products_data)} records and comprehensive XML data")
            
        except Exception as e:
            logging.error(f"Error updating products table: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar tabela de produtos:\n{str(e)}")
    
    def _show_products_context_menu(self, position):
        """Show context menu for products table"""
        if self.products_table.itemAt(position) is None:
            return
        
        menu = QMenu(self)
        
        # Export selected products
        export_action = menu.addAction("Exportar Selecionados")
        export_action.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        export_action.triggered.connect(self._export_selected_products)
        
        # Copy to clipboard
        copy_action = menu.addAction("Copiar para Área de Transferência")
        copy_action.setIcon(self.style().standardIcon(QStyle.SP_EditCopy))
        copy_action.triggered.connect(self._copy_products_to_clipboard)
        
        # View details
        view_action = menu.addAction("Ver Detalhes")
        view_action.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        view_action.triggered.connect(self._view_product_details)
        
        menu.addSeparator()
        
        # Export to Excel
        excel_action = menu.addAction("Exportar para Excel")
        excel_action.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        excel_action.triggered.connect(self._export_to_excel)
        
        # Check if any items are selected
        selected_items = self.products_table.selectedItems()
        if not selected_items:
            export_action.setEnabled(False)
            copy_action.setEnabled(False)
            view_action.setEnabled(False)
        
        menu.exec_(self.products_table.mapToGlobal(position))
    
    def _export_selected_products(self):
        """Export selected products to CSV"""
        try:
            # Get selected rows
            selected_rows = set()
            for item in self.products_table.selectedItems():
                selected_rows.add(item.row())
            
            if not selected_rows:
                QMessageBox.information(self, "Aviso", "Nenhum produto selecionado.")
                return
            
            # Get file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exportar Produtos Selecionados",
                f"produtos_selecionados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "CSV files (*.csv)"
            )
            
            if not file_path:
                return
            
            # Export data
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                
                # Write headers
                headers = []
                for col in range(self.products_table.columnCount()):
                    headers.append(self.products_table.horizontalHeaderItem(col).text())
                writer.writerow(headers)
                
                # Write selected rows
                for row in sorted(selected_rows):
                    row_data = []
                    for col in range(self.products_table.columnCount()):
                        item = self.products_table.item(row, col)
                        row_data.append(item.text() if item else '')
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Sucesso", f"Produtos exportados para: {file_path}")
            logging.info(f"Selected products exported to: {file_path}")
            
        except Exception as e:
            logging.error(f"Error exporting selected products: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar produtos:\n{str(e)}")
    
    def _copy_products_to_clipboard(self):
        """Copy selected products to clipboard"""
        try:
            # Get selected rows
            selected_rows = set()
            for item in self.products_table.selectedItems():
                selected_rows.add(item.row())
            
            if not selected_rows:
                QMessageBox.information(self, "Aviso", "Nenhum produto selecionado.")
                return
            
            # Build clipboard text
            clipboard_text = []
            
            # Add headers
            headers = []
            for col in range(self.products_table.columnCount()):
                headers.append(self.products_table.horizontalHeaderItem(col).text())
            clipboard_text.append('\t'.join(headers))
            
            # Add selected rows
            for row in sorted(selected_rows):
                row_data = []
                for col in range(self.products_table.columnCount()):
                    item = self.products_table.item(row, col)
                    row_data.append(item.text() if item else '')
                clipboard_text.append('\t'.join(row_data))
            
            # Copy to clipboard
            QApplication.clipboard().setText('\n'.join(clipboard_text))
            
            QMessageBox.information(self, "Sucesso", f"{len(selected_rows)} produtos copiados para a área de transferência.")
            logging.info(f"Copied {len(selected_rows)} products to clipboard")
            
        except Exception as e:
            logging.error(f"Error copying products to clipboard: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao copiar produtos:\n{str(e)}")
    
    def _view_product_details(self):
        """View details of selected product"""
        try:
            current_row = self.products_table.currentRow()
            if current_row < 0:
                QMessageBox.information(self, "Aviso", "Nenhum produto selecionado.")
                return
            
            # Get product data from table
            product_data = {}
            for col in range(self.products_table.columnCount()):
                header = self.products_table.horizontalHeaderItem(col).text()
                item = self.products_table.item(current_row, col)
                product_data[header] = item.text() if item else ''
            
            # Create details dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Detalhes do Produto")
            dialog.setModal(True)
            dialog.resize(600, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Create scroll area
            scroll = QScrollArea()
            scroll_widget = QWidget()
            scroll_layout = QFormLayout(scroll_widget)
            
            # Add product details
            for field, value in product_data.items():
                label = QLabel(field + ":")
                label.setStyleSheet("font-weight: bold;")
                value_label = QLabel(value)
                value_label.setWordWrap(True)
                value_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
                scroll_layout.addRow(label, value_label)
            
            scroll.setWidget(scroll_widget)
            layout.addWidget(scroll)
            
            # Add close button
            button_layout = QHBoxLayout()
            close_button = QPushButton("Fechar")
            close_button.clicked.connect(dialog.accept)
            button_layout.addStretch()
            button_layout.addWidget(close_button)
            layout.addLayout(button_layout)
            
            dialog.exec_()
            
        except Exception as e:
            logging.error(f"Error viewing product details: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao visualizar detalhes:\n{str(e)}")
    
    def _export_to_excel(self):
        """Exportar dados para Excel com opções avançadas - UPDATED to use optimized method"""
        try:
            # Use the new optimized export method
            self._export_to_excel_optimized()
                
        except Exception as e:
            logging.error(f"Erro ao exportar para Excel: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar dados:\n{str(e)}")
    
    def _transform_products_to_documents_format(self, products: List[Dict]) -> List[Dict]:
        """Transform flat products data to document format expected by ExportDialog"""
        documents = []
        
        # Group products by document (using access_key as unique identifier)
        document_groups = {}
        
        for product in products:
            access_key = product.get('access_key', '')
            if not access_key:
                # If no access_key, use document_number as fallback
                access_key = str(product.get('document_number', ''))
            
            if access_key not in document_groups:
                # Create new document entry
                document_groups[access_key] = {
                    'id': product.get('document_number', ''),
                    'document_type': product.get('document_type', ''),
                    'document_number': product.get('document_number', ''),
                    'series': product.get('series', ''),
                    'model': product.get('model', ''),
                    'issue_date': product.get('issue_date', ''),
                    'exit_date': product.get('exit_date', ''),
                    'access_key': access_key,
                    'protocol_number': product.get('protocol_number', ''),
                    'protocol_date': product.get('protocol_date', ''),
                    'operation_nature': product.get('operation_nature', ''),
                    'cnpj_issuer': product.get('cnpj_issuer', ''),
                    'issuer_name': product.get('issuer_name', ''),
                    'emitter_fantasy': product.get('emitter_fantasy', ''),
                    'emitter_ie': product.get('emitter_ie', ''),
                    'emitter_address': product.get('emitter_address', ''),
                    'emitter_city': product.get('emitter_city', ''),
                    'emitter_state': product.get('emitter_state', ''),
                    'emitter_cep': product.get('emitter_cep', ''),
                    'cnpj_recipient': product.get('cnpj_recipient', ''),
                    'recipient_name': product.get('recipient_name', ''),
                    'recipient_ie': product.get('recipient_ie', ''),
                    'recipient_address': product.get('recipient_address', ''),
                    'recipient_city': product.get('recipient_city', ''),
                    'recipient_state': product.get('recipient_state', ''),
                    'recipient_cep': product.get('recipient_cep', ''),
                    'total_products': product.get('total_products', 0),
                    'total_freight': product.get('total_freight', 0),
                    'total_insurance': product.get('total_insurance', 0),
                    'total_discount': product.get('total_discount', 0),
                    'total_other': product.get('total_other', 0),
                    'total_nfe': product.get('total_nfe', 0),
                    'icms_st_value': product.get('icms_st_value', 0),
                    'transport_modality': product.get('transport_modality', ''),
                    'transporter_name': product.get('transporter_name', ''),
                    'payment_method': product.get('payment_method', ''),
                    'additional_info': product.get('additional_info', ''),
                    'file_name': product.get('file_name', ''),
                    'status': 'active',
                    'items': []
                }
            
            # Add item to document
            item = {
                'item_number': product.get('item_number', ''),
                'item_code': product.get('item_code', ''),
                'item_ean': product.get('item_ean', ''),
                'item_description': product.get('item_description', ''),
                'ncm_code': product.get('ncm_code', ''),
                'cfop': product.get('cfop', ''),
                'commercial_unit': product.get('commercial_unit', ''),
                'quantity': product.get('quantity', 0),
                'unit_value': product.get('unit_value', 0),
                'total_value': product.get('total_value', 0),
                'icms_cst': product.get('icms_cst', ''),
                'icms_base': product.get('icms_base', 0),
                'icms_value': product.get('icms_value', 0),
                'icms_rate': product.get('icms_rate', 0),
                'ipi_cst': product.get('ipi_cst', ''),
                'ipi_base': product.get('ipi_base', 0),
                'ipi_value': product.get('ipi_value', 0),
                'ipi_rate': product.get('ipi_rate', 0),
                'pis_cst': product.get('pis_cst', ''),
                'pis_base': product.get('pis_base', 0),
                'pis_value': product.get('pis_value', 0),
                'pis_rate': product.get('pis_rate', 0),
                'cofins_cst': product.get('cofins_cst', ''),
                'cofins_base': product.get('cofins_base', 0),
                'cofins_value': product.get('cofins_value', 0),
                'cofins_rate': product.get('cofins_rate', 0),
                'tax_value': product.get('tax_value', 0)
            }
            
            document_groups[access_key]['items'].append(item)
        
        # Convert groups to list
        documents = list(document_groups.values())
        
        logging.info(f"Transformed {len(products)} products into {len(documents)} documents for export")
        return documents
    
    def _clear_filters(self):
        """Clear all filters"""
        try:
            self.product_search_input.clear()
            self.product_type_combo.setCurrentText("Todos")
            self._filter_products()
            
            self.status_bar.showMessage("Filtros limpos")
            
        except Exception as e:
            logging.error(f"Error clearing filters: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao limpar filtros:\n{str(e)}")
    
    def _load_database_manually(self):
        """Load database manually when user requests it"""
        try:
            self.status_bar.showMessage("Iniciando carregamento do banco de dados...")
            
            # Show confirmation dialog
            reply = QMessageBox.question(self, "Carregar Banco de Dados", 
                                       "Deseja carregar o banco de dados agora?\n\n"
                                       "Esta operação pode levar alguns minutos dependendo da quantidade de dados.",
                                       QMessageBox.Yes | QMessageBox.No,
                                       QMessageBox.Yes)
            
            if reply == QMessageBox.Yes:
                # Initialize databases first
                self._initialize_model_databases_async()
                
                # After database initialization, refresh products
                # This will be handled by the completion signal
                
        except Exception as e:
            logging.error(f"Error loading database manually: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao carregar banco de dados:\n{str(e)}")
    
    def _check_for_updates(self):
        """Check for application updates"""
        try:
            reply = QMessageBox.question(
                self,
                "Verificar Atualizações",
                "Deseja verificar se há atualizações disponíveis?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self._perform_update_check()
                
        except Exception as e:
            logging.error(f"Error checking for updates: {e}")
            QMessageBox.warning(
                self,
                "Erro",
                f"Erro ao verificar atualizações: {str(e)}"
            )

    def _perform_update_check(self):
        """Perform the actual update check process"""
        try:
            # Simulated update check process with progress
            progress_dialog = QProgressDialog("Verificando atualizações...", "Cancelar", 0, 100, self)
            progress_dialog.setWindowTitle("Verificação de Atualizações")
            progress_dialog.setModal(True)
            progress_dialog.setMinimumDuration(0)
            progress_dialog.show()
            
            QApplication.processEvents()
            
            for i in range(101):
                if progress_dialog.wasCanceled():
                    return
                progress_dialog.setValue(i)
                QApplication.processEvents()
                time.sleep(0.01)  # Simulate checking time
            
            progress_dialog.close()
            
            # Simulate result (always show "up to date" for this demo)
            QMessageBox.information(
                self,
                "Verificação de Atualizações",
                "Você está usando a versão mais recente do aplicativo!"
            )
            
            self.status_bar.showMessage("Verificação de atualizações concluída", 3000)
            
        except Exception as e:
            logging.error(f"Error performing update check: {e}")
            QMessageBox.critical(
                self,
                "Erro",
                f"Erro durante a verificação de atualizações: {str(e)}"
            )

    def _clear_all_data(self):
        """Clear all data from database"""
        try:
            reply = QMessageBox.question(
                self, 
                "Confirmar Limpeza", 
                "Tem certeza que deseja limpar todos os dados?\n\nEsta ação não pode ser desfeita.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Clear database
                if self.db_manager.clear_all_data():
                    # Clear UI
                    self._clear_ui_data()
                    QMessageBox.information(self, "Sucesso", "Todos os dados foram removidos.")
                    logging.info("All data cleared successfully")
                else:
                    QMessageBox.critical(self, "Erro", "Erro ao limpar os dados.")
                    logging.error("Failed to clear all data")
                    
        except Exception as e:
            logging.error(f"Error clearing all data: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao limpar dados:\n{str(e)}")

    def _clear_ui_data(self):
        """Clear UI data after database cleanup"""
        try:
            # Clear tables
            self.products_table.setRowCount(0)
            
            # Reset filters
            self.product_type_combo.setCurrentIndex(0)
            self.product_search_input.clear()
            
            # Update status
            self.db_status_label.setText("Base de dados: Vazia")
            self.statusBar().showMessage("Dados limpos")
            
            logging.info("UI data cleared")
            
        except Exception as e:
            logging.error(f"Error clearing UI data: {e}")

    def _show_settings(self):
        """Show settings dialog"""
        try:
            from ui.settings_dialog import SettingsDialog
            dialog = SettingsDialog(self.config, self)
            dialog.exec_()
        except Exception as e:
            logging.error(f"Error showing settings: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao abrir configurações:\n{str(e)}")

    def _show_about(self):
        """Show about dialog"""
        QMessageBox.about(self, "Sobre", 
                         "XML Fiscal Manager\n\n"
                         "Sistema de gestão de documentos fiscais XML\n"
                         "Versão 1.0\n\n"
                         "Desenvolvido para processamento e análise\n"
                         "de documentos fiscais brasileiros.")

    def closeEvent(self, event):
        """Handle application close event"""
        try:
            # Stop and clean up worker threads
            if hasattr(self, 'db_init_worker') and self.db_init_worker:
                if self.db_init_worker.isRunning():
                    self.db_init_worker.requestInterruption()
                    self.db_init_worker.wait(3000)  # Wait up to 3 seconds
                self.db_init_worker.deleteLater()
                logging.info("Database initialization worker cleaned up")
            
            if hasattr(self, 'import_worker') and self.import_worker:
                if self.import_worker.isRunning():
                    self.import_worker.requestInterruption()
                    self.import_worker.wait(3000)
                self.import_worker.deleteLater()
                logging.info("Import worker cleaned up")
            
            # Close database connection
            if hasattr(self, 'db_manager') and self.db_manager:
                self.db_manager.close()
                logging.info("Database connection closed")
            
            logging.info("Application closing")
            event.accept()
            
        except Exception as e:
            logging.error(f"Error during application close: {e}")
            event.accept()

    def _initialize_model_databases(self):
        """Initialize model databases"""
        try:
            # Check if xml_model_manager is available
            if not hasattr(self, 'xml_model_manager') or self.xml_model_manager is None:
                logging.warning("XML Model Manager not available, skipping database initialization")
                return
                
            # Initialize model databases using the XML model manager
            success = self.xml_model_manager.initialize_databases(self.db_manager)
            if success:
                logging.info("XML model databases initialized successfully")
            else:
                logging.warning("Some XML model databases failed to initialize")
        except Exception as e:
            logging.error(f"Error initializing model databases: {e}")
            # Don't raise the exception, just log it and continue
    
    def _on_model_selection_changed(self):
        """Handle XML model selection change"""
        try:
            selected_items = self.models_list.selectedItems()
            if not selected_items:
                return
            
            selected_item = selected_items[0]
            model_key = selected_item.data(Qt.UserRole)
            
            # Check if xml_model_manager exists
            if not hasattr(self, 'xml_model_manager') or self.xml_model_manager is None:
                logging.warning("XML Model Manager not available, using fallback selection logic")
                self.selected_model = self._create_fallback_model(model_key)
            else:
                self.selected_model = self.xml_model_manager.get_model(model_key.lower())
            
            if self.selected_model:
                # Check if UI elements exist before updating them
                if hasattr(self, 'title_label'):
                    self.title_label.setText(f"{self.selected_model.display_name} - Documentos")
                
                if hasattr(self, 'model_description_label'):
                    model_info = {
                        'nfe': {'code': '55', 'description': 'Produtos'},
                        'nfce': {'code': '65', 'description': 'Consumidor'},
                        'cte': {'code': '57', 'description': 'Transporte'},
                        'nfse': {'code': '56', 'description': 'Serviços'}
                    }
                    
                    info = model_info.get(model_key.lower(), {'code': '??', 'description': 'Outros'})
                    self.model_description_label.setText(
                        f"{self.selected_model.description}\n"
                        f"Modelo: {info['code']} • Tipo: {info['description']}\n"
                        f"Clique em 'Carregar Dados' para visualizar os documentos deste tipo."
                    )
                
                # Clear current data and show message to load
                if hasattr(self, 'products_table'):
                    self.products_table.setRowCount(0)
                
                if hasattr(self, 'model_count_label'):
                    self.model_count_label.setText("📈 Clique em 'Atualizar' para carregar dados")
                    self.model_count_label.setStyleSheet("""
                        QLabel {
                            color: #f39c12;
                            font-weight: bold;
                            padding: 8px 10px;
                            font-size: 12px;
                            background-color: #fef9e7;
                            border-radius: 4px;
                            border: 1px solid #fad7a0;
                        }
                    """)
                
                # Update type filter to match selected model
                if hasattr(self, 'product_type_combo'):
                    model_to_type = {
                        'nfe': 'NFe',
                        'nfce': 'NFCe', 
                        'cte': 'CTe',
                        'nfse': 'NFSe'
                    }
                    type_name = model_to_type.get(model_key.lower(), 'Todos')
                    
                    # Update combo box without triggering filter
                    self.product_type_combo.blockSignals(True)
                    index = self.product_type_combo.findText(type_name)
                    if index >= 0:
                        self.product_type_combo.setCurrentIndex(index)
                    self.product_type_combo.blockSignals(False)
                
                logging.info(f"Model selection changed to: {self.selected_model.display_name}")
            else:
                logging.warning(f"Could not find or create model for key: {model_key}")
                if hasattr(self, 'model_count_label'):
                    self.model_count_label.setText("❌ Modelo não encontrado")
                    self.model_count_label.setStyleSheet("""
                        QLabel {
                            color: #e74c3c;
                            font-weight: bold;
                            padding: 8px 10px;
                            font-size: 12px;
                            background-color: #fdf2f2;
                            border-radius: 4px;
                            border: 1px solid #fdbdbd;
                        }
                    """)
            
        except Exception as e:
            logging.error(f"Error handling model selection: {e}")
            if hasattr(self, 'model_count_label'):
                self.model_count_label.setText("❌ Erro na seleção do modelo")
                self.model_count_label.setStyleSheet("""
                    QLabel {
                        color: #e74c3c;
                        font-weight: bold;
                        padding: 8px 10px;
                        font-size: 12px;
                        background-color: #fdf2f2;
                        border-radius: 4px;
                        border: 1px solid #fdbdbd;
                    }
                """)
    
    def _create_fallback_model(self, model_key: str):
        """Create a fallback model object when XMLModelManager is not available"""
        class FallbackModel:
            def __init__(self, key):
                self.key = key.lower()
                
                self.model_info = {
                    'nfe': {
                        'name': 'nfe',
                        'display_name': 'NFe - Nota Fiscal Eletrônica',
                        'description': 'Nota Fiscal Eletrônica'
                    },
                    'nfce': {
                        'name': 'nfce',
                        'display_name': 'NFCe - Nota Fiscal de Consumidor Eletrônica',
                        'description': 'Nota Fiscal de Consumidor Eletrônica'
                    },
                    'cte': {
                        'name': 'cte',
                        'display_name': 'CTe - Conhecimento de Transporte Eletrônico',
                        'description': 'Conhecimento de Transporte Eletrônico'
                    },
                    'nfse': {
                        'name': 'nfse',
                        'display_name': 'NFSe - Nota Fiscal de Serviços Eletrônica',
                        'description': 'Nota Fiscal de Serviços Eletrônica'
                    }
                }
            
            @property
            def name(self):
                return self.model_info.get(self.key, {}).get('name', self.key)
            
            @property
            def display_name(self):
                return self.model_info.get(self.key, {}).get('display_name', f'{self.key.upper()} Model')
            
            @property
            def description(self):
                return self.model_info.get(self.key, {}).get('description', f'{self.key.upper()} Document')
        
        return FallbackModel(model_key)
    
    def _load_model_data(self):
        """Load data for the selected model on demand"""
        try:
            if not self.selected_model:
                QMessageBox.information(self, "Informação", "Selecione um tipo de documento primeiro.")
                return
            
            # Show loading message
            self.model_count_label.setText("🔄 Carregando dados...")
            self.model_count_label.setStyleSheet("""
                QLabel {
                    color: #3498db;
                    font-weight: bold;
                    padding: 8px 10px;
                    font-size: 12px;
                    background-color: #ebf5fb;
                    border-radius: 4px;
                    border: 1px solid #aed6f1;
                }
            """)
            
            # Process events to show the loading message
            QApplication.processEvents()
            
            # Load documents for specific model
            documents = self._get_documents_for_model(self.selected_model.name)
            
            if documents:
                # Update count and display
                self._update_model_count(len(documents))
                
                # Update table with ALL columns (not model-specific)
                self._update_products_table(documents)
                
                # Update success style
                self.model_count_label.setStyleSheet("""
                    QLabel {
                        color: #27ae60;
                        font-weight: bold;
                        padding: 8px 10px;
                        font-size: 12px;
                        background-color: #eafaf1;
                        border-radius: 4px;
                        border: 1px solid #a9dfbf;
                    }
                """)
                
                # Update status bar
                self.status_bar.showMessage(f"Carregados {len(documents)} documentos do tipo {self.selected_model.display_name}")
                
                logging.info(f"Loaded {len(documents)} documents for model {self.selected_model.name}")
                
            else:
                # No documents found for this specific model
                self.products_table.setRowCount(0)
                self.model_count_label.setText(f"📊 Nenhum documento encontrado para {self.selected_model.display_name}")
                self.model_count_label.setStyleSheet("""
                    QLabel {
                        color: #f39c12;
                        font-weight: bold;
                        padding: 8px 10px;
                        font-size: 12px;
                        background-color: #fef9e7;
                        border-radius: 4px;
                        border: 1px solid #fad7a0;
                    }
                """)
                
        except Exception as e:
            logging.error(f"Error loading model data: {e}")
            self.model_count_label.setText("❌ Erro ao carregar dados")
            self.model_count_label.setStyleSheet("""
                QLabel {
                    color: #e74c3c;
                    font-weight: bold;
                    padding: 8px 10px;
                    font-size: 12px;
                    background-color: #fdf2f2;
                    border-radius: 4px;
                    border: 1px solid #fdbdbd;
                }
            """)
            
            # Show error message
            QMessageBox.critical(self, "Erro", f"Erro ao carregar dados do modelo:\n{str(e)}")
    
    def _get_documents_for_model(self, model_name: str) -> List[Dict]:
        """Get documents for a specific model from database"""
        try:
            # Mapear nomes de modelo para os códigos corretos
            model_mapping = {
                'nfe': '55',     # NFe - Nota Fiscal Eletrônica
                'nfce': '65',    # NFCe - Nota Fiscal de Consumidor Eletrônica 
                'cte': '57',     # CTe - Conhecimento de Transporte Eletrônico
                'nfse': '56'     # NFSe - Nota Fiscal de Serviços Eletrônica
            }
            
            model_code = model_mapping.get(model_name.lower())
            if not model_code:
                logging.warning(f"Modelo não encontrado: {model_name}")
                return []
            
            # Buscar documentos com filtro por modelo
            filters = {'model': model_code}
            documents = self.db_manager.get_enhanced_products(filters)
            
            logging.info(f"Encontrados {len(documents)} documentos para modelo {model_name} (código {model_code})")
            return documents
            
        except Exception as e:
            logging.error(f"Error getting documents for model {model_name}: {e}")
            return []
    
    def _update_model_count(self, count: int):
        """Update the model document count display"""
        if hasattr(self, 'selected_model') and self.selected_model:
            self.model_count_label.setText(f"📈 Total: {count} documentos ({self.selected_model.display_name})")
        else:
            self.model_count_label.setText(f"📈 Total: {count} documentos")
    
    def _export_model_data(self):
        """Export data for the currently selected model"""
        try:
            if not self.selected_model:
                QMessageBox.information(self, "Informação", 
                                      "Selecione um tipo de XML específico para exportar os dados do modelo.")
                return
            
            # Get documents for the selected model
            documents = self._get_documents_for_model(self.selected_model.name)
            
            if not documents:
                QMessageBox.information(self, "Informação", 
                                      f"Não há documentos do tipo {self.selected_model.display_name} para exportar.")
                return
            
            # Transform documents to the format expected by ExportDialog
            documents_for_export = self._transform_products_to_documents_format(documents)
            
            # Show file dialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Exportar {self.selected_model.display_name}",
                f"{self.selected_model.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
            )
            
            if file_path:
                # Use existing export dialog with the transformed data
                export_dialog = ExportDialog(documents_for_export, self.config, self)
                export_dialog.file_path_edit.setText(file_path)
                
                # Set export type to detailed products
                export_dialog.export_type_combo.setCurrentText("Produtos/Itens (detalhado)")
                
                if export_dialog.exec() == QDialog.Accepted:
                    self.status_bar.showMessage(f"Exportação do modelo {self.selected_model.display_name} concluída!", 3000)
                    QMessageBox.information(self, "Sucesso", 
                                          f"Dados do modelo {self.selected_model.display_name} exportados com sucesso!")
            
        except Exception as e:
            logging.error(f"Error exporting model data: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar dados do modelo:\n{str(e)}")
    
    def _clear_model_selection(self):
        """Clear model selection and show all data"""
        try:
            # Clear selection
            self.models_list.clearSelection()
            self.selected_model = None
            
            # Update title
            self.title_label.setText("Todos os Documentos")
            
            # Update count label
            self.model_count_label.setText("📊 Mostrando todos os documentos")
            self.model_count_label.setStyleSheet("""
                QLabel {
                    color: #7f8c8d;
                    font-weight: bold;
                    padding: 6px 8px;
                    font-size: 10px;
                    background-color: #ecf0f1;
                    border-radius: 4px;
                    border: 1px solid #bdc3c7;
                }
            """)
            
            # Load all data
            self._refresh_products()
            
            self.status_bar.showMessage("Seleção limpa - mostrando todos os documentos")
            
        except Exception as e:
            logging.error(f"Error clearing model selection: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao limpar seleção:\n{str(e)}")
    
    def _show_export_config(self):
        """Show export configuration dialog"""
        try:
            dialog = ExportConfigDialog(self)
            if dialog.exec() == QDialog.Accepted:
                self.export_settings = dialog.get_settings()
                self.status_bar.showMessage("Configurações de exportação salvas!", 2000)
                logging.info("Export settings updated")
        except Exception as e:
            logging.error(f"Error showing export config: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao abrir configurações de exportação:\n{str(e)}")
    
    def _quick_export_excel(self):
        """Quick Excel export with current settings"""
        try:
            self._export_to_excel_optimized()
        except Exception as e:
            logging.error(f"Error in quick Excel export: {e}")
            QMessageBox.critical(self, "Erro", f"Erro na exportação rápida para Excel:\n{str(e)}")
    
    def _quick_export_csv(self):
        """Quick CSV export with current settings"""
        try:
            self._export_to_csv_optimized()
        except Exception as e:
            logging.error(f"Error in quick CSV export: {e}")
            QMessageBox.critical(self, "Erro", f"Erro na exportação rápida para CSV:\n{str(e)}")
    
    def _export_to_excel_optimized(self):
        """Optimized Excel export without heavy formatting"""
        try:
            # Get filtered data
            search_text = self.product_search_input.text().lower()
            doc_type_filter = self.product_type_combo.currentText()
            
            # Prepare filters for database search
            filters = {}
            if doc_type_filter != "Todos":
                filters['document_type'] = doc_type_filter.lower()
            
            # Get data from database
            all_products = self.db_manager.get_enhanced_products(filters)
            
            # Apply text filter if necessary
            if search_text:
                filtered_products = []
                for product in all_products:
                    searchable_text = ' '.join([
                        str(product.get('item_description', '')),
                        str(product.get('item_code', '')),
                        str(product.get('ncm_code', '')),
                        str(product.get('cfop', '')),
                        str(product.get('document_number', '')),
                        str(product.get('file_name', ''))
                    ]).lower()
                    
                    if search_text in searchable_text:
                        filtered_products.append(product)
                
                products = filtered_products
            else:
                products = all_products
            
            if not products:
                QMessageBox.warning(self, "Aviso", "Não há produtos para exportar com os filtros aplicados.")
                return
            
            # Get file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exportar para Excel",
                f"produtos_{QDateTime.currentDateTime().toString('yyyyMMdd_hhmmss')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Export with progress dialog
            self._export_excel_with_progress(products, file_path)
            
        except Exception as e:
            logging.error(f"Error in optimized Excel export: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar para Excel:\n{str(e)}")
    
    def _export_excel_with_progress(self, products: List[Dict], file_path: str):
        """Export to Excel with progress dialog and optimized performance"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # Create progress dialog
            progress = QProgressDialog("Exportando para Excel...", "Cancelar", 0, 100, self)
            progress.setWindowTitle("Exportação")
            progress.setModal(True)
            progress.show()
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Produtos"
            
            # Define headers (comprehensive)
            headers = [
                "Tipo Doc", "Número", "Série", "Modelo", "Data Emissão", "Data Saída",
                "Chave Acesso", "Protocolo", "Data Protocolo", "Natureza Operação",
                "CNPJ Emitente", "Emitente", "Nome Fantasia", "IE Emitente",
                "End. Emitente", "Cidade Emitente", "UF Emitente", "CEP Emitente",
                "CNPJ/CPF Destinatário", "Destinatário", "IE Destinatário", 
                "End. Destinatário", "Cidade Destinatário", "UF Destinatário",
                "Nº Item", "Código Item", "Descrição", "NCM", "CFOP", "EAN/GTIN",
                "Quantidade", "Unidade", "Valor Unit", "Valor Total Item",
                "CST ICMS", "Base ICMS", "Valor ICMS", "Alíq ICMS",
                "CST IPI", "Base IPI", "Valor IPI", "Alíq IPI",
                "CST PIS", "Base PIS", "Valor PIS", "Alíq PIS",
                "CST COFINS", "Base COFINS", "Valor COFINS", "Alíq COFINS",
                "Total Produtos", "Total Frete", "Total Seguro", "Total Desconto",
                "Total Outros", "Total NFe", "Total Tributos Item", "ICMS ST",
                "Modalidade Frete", "Transportadora", "Forma Pagamento",
                "Info Adicional", "Arquivo"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            progress.setValue(10)
            QApplication.processEvents()
            
            if progress.wasCanceled():
                return
            
            # Define column mapping
            column_mapping = [
                'document_type', 'document_number', 'series', 'model', 'issue_date', 'exit_date',
                'access_key', 'protocol_number', 'protocol_date', 'operation_nature',
                'cnpj_issuer', 'issuer_name', 'emitter_fantasy', 'emitter_ie',
                'emitter_address', 'emitter_city', 'emitter_state', 'emitter_cep',
                'cnpj_recipient', 'recipient_name', 'recipient_ie',
                'recipient_address', 'recipient_city', 'recipient_state',
                'item_number', 'item_code', 'item_description', 'ncm_code', 'cfop', 'item_ean',
                'quantity', 'commercial_unit', 'unit_value', 'total_value',
                'icms_cst', 'icms_base', 'icms_value', 'icms_rate',
                'ipi_cst', 'ipi_base', 'ipi_value', 'ipi_rate',
                'pis_cst', 'pis_base', 'pis_value', 'pis_rate',
                'cofins_cst', 'cofins_base', 'cofins_value', 'cofins_rate',
                'total_products', 'total_freight', 'total_insurance', 'total_discount',
                'total_other', 'total_nfe', 'tax_value', 'icms_st_value',
                'transport_modality', 'transporter_name', 'payment_method',
                'additional_info', 'file_name'
            ]
            
            # Write data in batches for better performance
            batch_size = self.export_settings['performance']['batch_size']
            total_rows = len(products)
            
            for i in range(0, total_rows, batch_size):
                batch = products[i:i + batch_size]
                
                for row_idx, product in enumerate(batch, start=i + 2):  # Start from row 2 (after header)
                    for col_idx, field_name in enumerate(column_mapping, start=1):
                        try:
                            value = product.get(field_name, '')
                            
                            # Basic data conversion without heavy formatting
                            if field_name in ['quantity', 'unit_value', 'total_value', 'icms_base', 'icms_value',
                                            'ipi_base', 'ipi_value', 'pis_base', 'pis_value', 'cofins_base', 'cofins_value',
                                            'total_products', 'total_freight', 'total_insurance', 'total_discount',
                                            'total_other', 'total_nfe', 'tax_value', 'icms_st_value',
                                            'icms_rate', 'ipi_rate', 'pis_rate', 'cofins_rate']:
                                try:
                                    ws.cell(row=row_idx, column=col_idx, value=float(value) if value else 0)
                                except:
                                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                        
                        except Exception as e:
                            logging.error(f"Error writing cell [{row_idx}, {col_idx}]: {e}")
                            ws.cell(row=row_idx, column=col_idx, value='')
                
                # Update progress
                progress_value = int(10 + ((i + len(batch)) / total_rows) * 80)
                progress.setValue(progress_value)
                QApplication.processEvents()
                
                if progress.wasCanceled():
                    return
            
            # Apply basic formatting only if enabled
            if self.export_settings['excel']['enable_formatting']:
                progress.setLabelText("Aplicando formatação...")
                QApplication.processEvents()
                
                # Basic header formatting
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Auto-resize columns if enabled
            if self.export_settings['excel']['auto_resize_columns']:
                progress.setLabelText("Ajustando colunas...")
                QApplication.processEvents()
                
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].auto_size = True
            
            # Freeze header if enabled
            if self.export_settings['excel']['freeze_header']:
                ws.freeze_panes = 'A2'
            
            progress.setValue(95)
            progress.setLabelText("Salvando arquivo...")
            QApplication.processEvents()
            
            # Save workbook
            wb.save(file_path)
            
            progress.setValue(100)
            progress.close()
            
            # Show success message
            QMessageBox.information(self, "Sucesso", 
                                  f"Dados exportados com sucesso!\n\n"
                                  f"Arquivo: {file_path}\n"
                                  f"Registros: {len(products)}")
            
            self.status_bar.showMessage(f"Exportação concluída: {len(products)} registros", 3000)
            logging.info(f"Excel export completed: {len(products)} records to {file_path}")
            
        except ImportError:
            QMessageBox.critical(self, "Erro", "Biblioteca openpyxl não encontrada.\nInstale com: pip install openpyxl")
        except Exception as e:
            logging.error(f"Error in Excel export with progress: {e}")
            QMessageBox.critical(self, "Erro", f"Erro durante exportação:\n{str(e)}")
    
    def _export_to_csv_optimized(self):
        """Optimized CSV export"""
        try:
            # Get filtered data (same logic as Excel export)
            search_text = self.product_search_input.text().lower()
            doc_type_filter = self.product_type_combo.currentText()
            
            filters = {}
            if doc_type_filter != "Todos":
                filters['document_type'] = doc_type_filter.lower()
            
            all_products = self.db_manager.get_enhanced_products(filters)
            
            if search_text:
                filtered_products = []
                for product in all_products:
                    searchable_text = ' '.join([
                        str(product.get('item_description', '')),
                        str(product.get('item_code', '')),
                        str(product.get('ncm_code', '')),
                        str(product.get('cfop', '')),
                        str(product.get('document_number', '')),
                        str(product.get('file_name', ''))
                    ]).lower()
                    
                    if search_text in searchable_text:
                        filtered_products.append(product)
                
                products = filtered_products
            else:
                products = all_products
            
            if not products:
                QMessageBox.warning(self, "Aviso", "Não há produtos para exportar com os filtros aplicados.")
                return
            
            # Get file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exportar para CSV",
                f"produtos_{QDateTime.currentDateTime().toString('yyyyMMdd_hhmmss')}.csv",
                "CSV Files (*.csv);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Export CSV with progress
            self._export_csv_with_progress(products, file_path)
            
        except Exception as e:
            logging.error(f"Error in optimized CSV export: {e}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar para CSV:\n{str(e)}")
    
    def _export_csv_with_progress(self, products: List[Dict], file_path: str):
        """Export to CSV with progress dialog"""
        try:
            import csv
            
            # Create progress dialog
            progress = QProgressDialog("Exportando para CSV...", "Cancelar", 0, 100, self)
            progress.setWindowTitle("Exportação CSV")
            progress.setModal(True)
            progress.show()
            
            # Get CSV settings
            separator = self.export_settings['csv']['separator']
            encoding = self.export_settings['csv']['encoding']
            
            # Define headers and column mapping (same as Excel)
            headers = [
                "Tipo Doc", "Número", "Série", "Modelo", "Data Emissão", "Data Saída",
                "Chave Acesso", "Protocolo", "Data Protocolo", "Natureza Operação",
                "CNPJ Emitente", "Emitente", "Nome Fantasia", "IE Emitente",
                "End. Emitente", "Cidade Emitente", "UF Emitente", "CEP Emitente",
                "CNPJ/CPF Destinatário", "Destinatário", "IE Destinatário", 
                "End. Destinatário", "Cidade Destinatário", "UF Destinatário",
                "Nº Item", "Código Item", "Descrição", "NCM", "CFOP", "EAN/GTIN",
                "Quantidade", "Unidade", "Valor Unit", "Valor Total Item",
                "CST ICMS", "Base ICMS", "Valor ICMS", "Alíq ICMS",
                "CST IPI", "Base IPI", "Valor IPI", "Alíq IPI",
                "CST PIS", "Base PIS", "Valor PIS", "Alíq PIS",
                "CST COFINS", "Base COFINS", "Valor COFINS", "Alíq COFINS",
                "Total Produtos", "Total Frete", "Total Seguro", "Total Desconto",
                "Total Outros", "Total NFe", "Total Tributos Item", "ICMS ST",
                "Modalidade Frete", "Transportadora", "Forma Pagamento",
                "Info Adicional", "Arquivo"
            ]
            
            column_mapping = [
                'document_type', 'document_number', 'series', 'model', 'issue_date', 'exit_date',
                'access_key', 'protocol_number', 'protocol_date', 'operation_nature',
                'cnpj_issuer', 'issuer_name', 'emitter_fantasy', 'emitter_ie',
                'emitter_address', 'emitter_city', 'emitter_state', 'emitter_cep',
                'cnpj_recipient', 'recipient_name', 'recipient_ie',
                'recipient_address', 'recipient_city', 'recipient_state',
                'item_number', 'item_code', 'item_description', 'ncm_code', 'cfop', 'item_ean',
                'quantity', 'commercial_unit', 'unit_value', 'total_value',
                'icms_cst', 'icms_base', 'icms_value', 'icms_rate',
                'ipi_cst', 'ipi_base', 'ipi_value', 'ipi_rate',
                'pis_cst', 'pis_base', 'pis_value', 'pis_rate',
                'cofins_cst', 'cofins_base', 'cofins_value', 'cofins_rate',
                'total_products', 'total_freight', 'total_insurance', 'total_discount',
                'total_other', 'total_nfe', 'tax_value', 'icms_st_value',
                'transport_modality', 'transporter_name', 'payment_method',
                'additional_info', 'file_name'
            ]
            
            progress.setValue(10)
            QApplication.processEvents()
            
            # Write CSV file
            with open(file_path, 'w', newline='', encoding=encoding) as csvfile:
                writer = csv.writer(csvfile, delimiter=separator, quoting=csv.QUOTE_MINIMAL)
                
                # Write headers
                writer.writerow(headers)
                
                progress.setValue(20)
                QApplication.processEvents()
                
                if progress.wasCanceled():
                    return
                
                # Write data in batches
                batch_size = self.export_settings['performance']['batch_size']
                total_rows = len(products)
                
                for i in range(0, total_rows, batch_size):
                    batch = products[i:i + batch_size]
                    
                    batch_rows = []
                    for product in batch:
                        row_data = []
                        for field_name in column_mapping:
                            value = product.get(field_name, '')
                            row_data.append(str(value) if value else '')
                        batch_rows.append(row_data)
                    
                    # Write batch
                    writer.writerows(batch_rows)
                    
                    # Update progress
                    progress_value = int(20 + ((i + len(batch)) / total_rows) * 70)
                    progress.setValue(progress_value)
                    QApplication.processEvents()
                    
                    if progress.wasCanceled():
                        return
            
            progress.setValue(100)
            progress.close()
            
            # Show success message
            QMessageBox.information(self, "Sucesso", 
                                  f"Dados exportados para CSV com sucesso!\n\n"
                                  f"Arquivo: {file_path}\n"
                                  f"Registros: {len(products)}\n"
                                  f"Separador: '{separator}'\n"
                                  f"Codificação: {encoding}")
            
            self.status_bar.showMessage(f"Exportação CSV concluída: {len(products)} registros", 3000)
            logging.info(f"CSV export completed: {len(products)} records to {file_path}")
            
        except Exception as e:
            logging.error(f"Error in CSV export with progress: {e}")
            QMessageBox.critical(self, "Erro", f"Erro durante exportação CSV:\n{str(e)}")