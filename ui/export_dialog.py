#!/usr/bin/env python3
"""
Export Dialog for XML Fiscal Manager Pro
Advanced export functionality with multiple formats and options
"""

import json
import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QCheckBox, QComboBox, QLineEdit, QFileDialog,
    QMessageBox, QProgressBar, QTextEdit, QGridLayout,
    QDialogButtonBox, QFrame, QSpinBox, QDateEdit, QTabWidget,
    QListWidget, QListWidgetItem, QFormLayout
)
from PySide6.QtCore import Qt, QThread, Signal, QDate
from PySide6.QtGui import QFont
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportWorker(QThread):
    """Worker thread for exporting documents"""
    
    progress_updated = Signal(int, str)
    finished = Signal(bool, str)
    error_occurred = Signal(str)
    
    def __init__(self, documents, export_config, output_path):
        super().__init__()
        self.documents = documents
        self.export_config = export_config
        self.output_path = output_path
    
    def run(self):
        """Export documents or products in background"""
        try:
            self.progress_updated.emit(0, "Iniciando exportação...")
            
            export_format = self.export_config['format']
            export_type = self.export_config.get('export_type', 'documents')  # 'documents' or 'products'
            
            if export_type == 'products':
                # Export products/items specifically
                if export_format == 'excel':
                    success, message = self._export_products_excel()
                elif export_format == 'csv':
                    success, message = self._export_products_csv()
                elif export_format == 'json':
                    success, message = self._export_products_json()
                else:
                    success, message = False, f"Formato não suportado para produtos: {export_format}"
            else:
                # Default document export
                if export_format == 'excel':
                    success, message = self._export_excel()
                elif export_format == 'csv':
                    success, message = self._export_csv()
                elif export_format == 'json':
                    success, message = self._export_json()
                else:
                    success, message = False, f"Formato não suportado: {export_format}"
            
            self.finished.emit(success, message)
            
        except Exception as e:
            self.error_occurred.emit(str(e))
    
    def _export_excel(self):
        """Export to Excel format"""
        if not EXCEL_AVAILABLE:
            # Try alternative export method
            return self._export_excel_alternative()
        
        try:
            self.progress_updated.emit(10, "Preparando dados para Excel...")
            
            # Prepare data
            data = self._prepare_data()
            if not data:
                return False, "Nenhum dado para exportar"
            
            df = pd.DataFrame(data)
            
            self.progress_updated.emit(30, "Criando arquivo Excel...")
            
            # Create workbook with error handling
            try:
                with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                    # Write main data
                    df.to_excel(writer, sheet_name='Documentos', index=False)
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Documentos']
                    
                    self.progress_updated.emit(60, "Formatando planilha...")
                    
                    # Apply formatting
                    self._format_excel_worksheet(worksheet, df)
                    
                    # Add summary sheet if configured
                    if self.export_config.get('include_summary', True):
                        self._add_summary_sheet(writer, df)
                    
                    self.progress_updated.emit(90, "Finalizando arquivo...")
                    
            except Exception as e:
                # If openpyxl fails, try xlsxwriter
                try:
                    with pd.ExcelWriter(self.output_path, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='Documentos', index=False)
                        self.progress_updated.emit(90, "Finalizando arquivo...")
                except Exception as e2:
                    return False, f"Erro ao criar arquivo Excel: {str(e2)}"
            
            self.progress_updated.emit(100, "Exportação concluída!")
            return True, f"Arquivo Excel criado: {self.output_path}"
            
        except Exception as e:
            logging.error(f"Excel export error: {e}")
            return False, f"Erro ao exportar Excel: {str(e)}"

    def _export_excel_alternative(self):
        """Alternative Excel export using CSV format when Excel libraries are not available"""
        try:
            self.progress_updated.emit(10, "Bibliotecas Excel não disponíveis, usando formato CSV...")
            
            # Change output path to CSV
            csv_path = str(self.output_path).replace('.xlsx', '.csv').replace('.xls', '.csv')
            
            data = self._prepare_data()
            if not data:
                return False, "Nenhum dado para exportar"
            
            self.progress_updated.emit(50, "Escrevendo arquivo CSV...")
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
                    
                    writer.writeheader()
                    
                    for i, row in enumerate(data):
                        writer.writerow(row)
                        if i % 100 == 0:
                            progress = 50 + int((i / len(data)) * 40)
                            self.progress_updated.emit(progress, f"Processando linha {i+1}...")
            
            self.progress_updated.emit(100, "Exportação concluída!")
            return True, f"Arquivo CSV criado (Excel não disponível): {csv_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar CSV alternativo: {str(e)}"
    
    def _export_csv(self):
        """Export to CSV format"""
        try:
            self.progress_updated.emit(20, "Preparando dados para CSV...")
            
            data = self._prepare_data()
            
            self.progress_updated.emit(50, "Escrevendo arquivo CSV...")
            
            with open(self.output_path, 'w', newline='', encoding='utf-8') as csvfile:
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, 
                                          delimiter=self.export_config.get('csv_delimiter', ','))
                    
                    if self.export_config.get('include_header', True):
                        writer.writeheader()
                    
                    for i, row in enumerate(data):
                        writer.writerow(row)
                        if i % 100 == 0:
                            progress = 50 + int((i / len(data)) * 40)
                            self.progress_updated.emit(progress, f"Processando linha {i+1}...")
            
            self.progress_updated.emit(100, "Exportação concluída!")
            return True, f"Arquivo CSV criado: {self.output_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar CSV: {str(e)}"
    
    def _export_json(self):
        """Export to JSON format"""
        try:
            self.progress_updated.emit(20, "Preparando dados para JSON...")
            
            data = self._prepare_data()
            
            self.progress_updated.emit(50, "Escrevendo arquivo JSON...")
            
            # Prepare JSON structure
            export_data = {
                'export_info': {
                    'timestamp': datetime.now().isoformat(),
                    'total_documents': len(data),
                    'format_version': '1.0'
                },
                'documents': data
            }
            
            with open(self.output_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)
            
            self.progress_updated.emit(100, "Exportação concluída!")
            return True, f"Arquivo JSON criado: {self.output_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar JSON: {str(e)}"
    
    def _prepare_data(self):
        """Prepare document data for export with comprehensive Brazilian XML tags"""
        data = []
        include_fields = self.export_config.get('include_fields', [])
        date_format = self.export_config.get('date_format', '%d/%m/%Y')
        
        for doc in self.documents:
            try:
                row = {}
                
                # === IDENTIFICAÇÃO DO DOCUMENTO ===
                row.update({
                    'ID_Interno': str(doc.get('id', '')),
                    'Chave_Acesso': str(doc.get('access_key', '')),
                    'Tipo_Documento': str(doc.get('document_type', '')).upper(),
                    'Modelo_Documento': str(doc.get('model', '')),
                    'Serie_Documento': self._format_number_or_text(doc.get('series', '')),
                    'Numero_Documento': self._format_number_or_text(doc.get('document_number', '')),
                    'Versao_Documento': str(doc.get('version', '')),
                    'Finalidade_NFe': str(doc.get('finalidade_nfe', '')),
                    'Processo_Emissao': str(doc.get('processo_emissao', '')),
                    'Versao_Processo': str(doc.get('versao_processo', ''))
                })
                
                # === DATAS ===
                row.update({
                    'Data_Emissao': self._format_date(doc.get('issue_date'), date_format),
                    'Data_Entrada_Saida': self._format_date(doc.get('entry_exit_date'), date_format),
                    'Data_Processamento': self._format_date(doc.get('processed_date'), date_format),
                    'Data_Criacao_Sistema': self._format_date(doc.get('created_at'), date_format),
                    'Data_Ultima_Atualizacao': self._format_date(doc.get('updated_at'), date_format),
                    'Hora_Entrada_Saida': str(doc.get('hora_saida', '')),
                    'Data_Contingencia': self._format_date(doc.get('data_contingencia'), date_format)
                })
                
                # === STATUS E SITUAÇÃO ===
                row.update({
                    'Status_Documento': str(doc.get('status', '')).title(),
                    'Situacao_Documento': str(doc.get('situacao', '')),
                    'Codigo_Status': str(doc.get('codigo_status', '')),
                    'Motivo_Status': str(doc.get('motivo_status', '')),
                    'Protocolo_Autorizacao': str(doc.get('protocol', '')),
                    'Data_Autorizacao': self._format_date(doc.get('data_autorizacao'), date_format),
                    'Justificativa_Cancelamento': str(doc.get('justification', ''))
                })
                
                # === EMITENTE - IDENTIFICAÇÃO ===
                row.update({
                    'CNPJ_Emitente': self._format_number_or_text(doc.get('cnpj_issuer', '')),
                    'CPF_Emitente': self._format_number_or_text(doc.get('cpf_issuer', '')),
                    'Razao_Social_Emitente': str(doc.get('issuer_name', '')),
                    'Nome_Fantasia_Emitente': str(doc.get('issuer_trade_name', '')),
                    'Inscricao_Estadual_Emitente': self._format_number_or_text(doc.get('issuer_state_registration', '')),
                    'Inscricao_Municipal_Emitente': self._format_number_or_text(doc.get('issuer_municipal_registration', '')),
                    'CNAE_Emitente': self._format_number_or_text(doc.get('cnae_emitente', '')),
                    'Regime_Tributario_Emitente': str(doc.get('regime_tributario_emitente', ''))
                })
                
                # === EMITENTE - ENDEREÇO ===
                row.update({
                    'Endereco_Emitente': str(doc.get('issuer_address', '')),
                    'Numero_Emitente': str(doc.get('issuer_number', '')),
                    'Complemento_Emitente': str(doc.get('issuer_complement', '')),
                    'Bairro_Emitente': str(doc.get('issuer_district', '')),
                    'Cidade_Emitente': str(doc.get('issuer_city', '')),
                    'UF_Emitente': str(doc.get('issuer_state', '')),
                    'CEP_Emitente': self._format_number_or_text(doc.get('issuer_zip_code', '')),
                    'Codigo_Municipio_Emitente': self._format_number_or_text(doc.get('cod_municipio_emitente', '')),
                    'Pais_Emitente': str(doc.get('pais_emitente', 'Brasil')),
                    'Codigo_Pais_Emitente': self._format_number_or_text(doc.get('cod_pais_emitente', '1058')),
                    'Telefone_Emitente': self._format_number_or_text(doc.get('issuer_phone', '')),
                    'Email_Emitente': str(doc.get('issuer_email', ''))
                })
                
                # === DESTINATÁRIO - IDENTIFICAÇÃO ===
                row.update({
                    'CNPJ_Destinatario': self._format_number_or_text(doc.get('cnpj_recipient', '')),
                    'CPF_Destinatario': self._format_number_or_text(doc.get('cpf_recipient', '')),
                    'Razao_Social_Destinatario': str(doc.get('recipient_name', '')),
                    'Nome_Fantasia_Destinatario': str(doc.get('recipient_trade_name', '')),
                    'Inscricao_Estadual_Destinatario': self._format_number_or_text(doc.get('recipient_state_registration', '')),
                    'Inscricao_Municipal_Destinatario': self._format_number_or_text(doc.get('recipient_municipal_registration', '')),
                    'Inscricao_SUFRAMA': self._format_number_or_text(doc.get('inscricao_suframa', '')),
                    'Indicador_IE_Destinatario': str(doc.get('indicador_ie_dest', ''))
                })
                
                # === DESTINATÁRIO - ENDEREÇO ===
                row.update({
                    'Endereco_Destinatario': str(doc.get('recipient_address', '')),
                    'Numero_Destinatario': str(doc.get('recipient_number', '')),
                    'Complemento_Destinatario': str(doc.get('recipient_complement', '')),
                    'Bairro_Destinatario': str(doc.get('recipient_district', '')),
                    'Cidade_Destinatario': str(doc.get('recipient_city', '')),
                    'UF_Destinatario': str(doc.get('recipient_state', '')),
                    'CEP_Destinatario': self._format_number_or_text(doc.get('recipient_zip_code', '')),
                    'Codigo_Municipio_Destinatario': self._format_number_or_text(doc.get('cod_municipio_destinatario', '')),
                    'Pais_Destinatario': str(doc.get('pais_destinatario', 'Brasil')),
                    'Codigo_Pais_Destinatario': self._format_number_or_text(doc.get('cod_pais_destinatario', '1058')),
                    'Telefone_Destinatario': self._format_number_or_text(doc.get('recipient_phone', '')),
                    'Email_Destinatario': str(doc.get('recipient_email', ''))
                })
                
                # === OPERAÇÃO ===
                row.update({
                    'Natureza_Operacao': str(doc.get('operation_nature', '')),
                    'CFOP_Operacao': self._format_number_or_text(doc.get('cfop_operacao', '')),
                    'Tipo_Operacao': str(doc.get('tipo_operacao', '')),
                    'Indicador_Presenca': str(doc.get('indicador_presenca', '')),
                    'Indicador_Consumidor_Final': str(doc.get('indicador_consumidor_final', '')),
                    'Local_Destino': str(doc.get('local_destino', '')),
                    'Municipio_Ocorrencia_Fato': str(doc.get('municipio_fato_gerador', '')),
                    'Tipo_Impressao_DANFE': str(doc.get('tipo_impressao_danfe', ''))
                })
                
                # === VALORES TOTAIS ===
                row.update({
                    'Valor_Total_NFe': self._format_decimal(doc.get('total_value', 0)),
                    'Valor_Total_Produtos': self._format_decimal(doc.get('products_value', 0)),
                    'Valor_Total_Servicos': self._format_decimal(doc.get('services_value', 0)),
                    'Valor_Total_Desconto': self._format_decimal(doc.get('discount_value', 0)),
                    'Valor_Total_Acrescimos': self._format_decimal(doc.get('acrescimos_value', 0)),
                    'Valor_Frete': self._format_decimal(doc.get('freight_value', 0)),
                    'Valor_Seguro': self._format_decimal(doc.get('insurance_value', 0)),
                    'Outras_Despesas_Acessorias': self._format_decimal(doc.get('other_expenses', 0)),
                    'Valor_Total_II': self._format_decimal(doc.get('ii_value', 0)),
                    'Valor_IOF': self._format_decimal(doc.get('iof_value', 0))
                })
                
                # === ICMS ===
                row.update({
                    'Base_Calculo_ICMS': self._format_decimal(doc.get('icms_base', 0)),
                    'Valor_ICMS': self._format_decimal(doc.get('icms_value', 0)),
                    'Valor_ICMS_Desonerado': self._format_decimal(doc.get('icms_desonerado', 0)),
                    'Base_Calculo_ICMS_ST': self._format_decimal(doc.get('icms_st_base', 0)),
                    'Valor_ICMS_ST': self._format_decimal(doc.get('icms_st_value', 0)),
                    'Valor_Total_Produtos_ST': self._format_decimal(doc.get('produtos_st_value', 0)),
                    'Base_ICMS_FCP': self._format_decimal(doc.get('icms_fcp_base', 0)),
                    'Valor_ICMS_FCP': self._format_decimal(doc.get('icms_fcp_value', 0)),
                    'Base_ICMS_ST_FCP': self._format_decimal(doc.get('icms_st_fcp_base', 0)),
                    'Valor_ICMS_ST_FCP': self._format_decimal(doc.get('icms_st_fcp_value', 0)),
                    'Valor_Total_FCP': self._format_decimal(doc.get('fcp_total_value', 0))
                })
                
                # === IPI ===
                row.update({
                    'Valor_Total_IPI': self._format_decimal(doc.get('ipi_value', 0)),
                    'Valor_IPI_Devolvido': self._format_decimal(doc.get('ipi_devolvido', 0))
                })
                
                # === PIS ===
                row.update({
                    'Valor_Total_PIS': self._format_decimal(doc.get('pis_value', 0)),
                    'Base_Calculo_PIS': self._format_decimal(doc.get('pis_base', 0))
                })
                
                # === COFINS ===
                row.update({
                    'Valor_Total_COFINS': self._format_decimal(doc.get('cofins_value', 0)),
                    'Base_Calculo_COFINS': self._format_decimal(doc.get('cofins_base', 0))
                })
                
                # === OUTROS TRIBUTOS ===
                row.update({
                    'Valor_Total_Tributos': self._format_decimal(doc.get('tax_value', 0)),
                    'Valor_Total_ISSQN': self._format_decimal(doc.get('issqn_value', 0)),
                    'Base_Calculo_ISSQN': self._format_decimal(doc.get('issqn_base', 0)),
                    'Aliquota_ISSQN': self._format_decimal(doc.get('issqn_aliquota', 0)),
                    'Codigo_Servico_ISSQN': self._format_number_or_text(doc.get('cod_servico_issqn', '')),
                    'Codigo_Municipio_ISSQN': self._format_number_or_text(doc.get('cod_municipio_issqn', '')),
                    'Valor_Deducoes_ISSQN': self._format_decimal(doc.get('deducoes_issqn', 0)),
                    'Valor_Outras_Retencoes': self._format_decimal(doc.get('outras_retencoes', 0)),
                    'Valor_Desconto_Incondicionado': self._format_decimal(doc.get('desconto_incondicionado', 0)),
                    'Valor_Desconto_Condicionado': self._format_decimal(doc.get('desconto_condicionado', 0))
                })
                
                # === RETENÇÕES ===
                row.update({
                    'Valor_Retencao_PIS': self._format_decimal(doc.get('retencao_pis', 0)),
                    'Valor_Retencao_COFINS': self._format_decimal(doc.get('retencao_cofins', 0)),
                    'Valor_Retencao_CSLL': self._format_decimal(doc.get('retencao_csll', 0)),
                    'Valor_Retencao_IRRF': self._format_decimal(doc.get('retencao_irrf', 0)),
                    'Base_Calculo_IRRF': self._format_decimal(doc.get('base_irrf', 0)),
                    'Valor_Retencao_Previdencia': self._format_decimal(doc.get('retencao_previdencia', 0)),
                    'Base_Calculo_Previdencia': self._format_decimal(doc.get('base_previdencia', 0))
                })
                
                # === TRANSPORTE ===
                row.update({
                    'Modalidade_Frete': str(doc.get('freight_modality', '')),
                    'CNPJ_CPF_Transportadora': self._format_number_or_text(doc.get('carrier_cnpj', '')),
                    'Razao_Social_Transportadora': str(doc.get('carrier_name', '')),
                    'Inscricao_Estadual_Transportadora': self._format_number_or_text(doc.get('carrier_ie', '')),
                    'Endereco_Transportadora': str(doc.get('carrier_address', '')),
                    'Municipio_Transportadora': str(doc.get('carrier_city', '')),
                    'UF_Transportadora': str(doc.get('carrier_state', '')),
                    'Placa_Veiculo': str(doc.get('vehicle_plate', '')),
                    'UF_Veiculo': str(doc.get('vehicle_state', '')),
                    'RNTC_Veiculo': str(doc.get('vehicle_rntc', ''))
                })
                
                # === VOLUMES ===
                row.update({
                    'Quantidade_Volumes': self._format_decimal(doc.get('volumes_quantity', 0)),
                    'Especie_Volumes': str(doc.get('volumes_species', '')),
                    'Marca_Volumes': str(doc.get('volumes_brand', '')),
                    'Numeracao_Volumes': str(doc.get('volumes_number', '')),
                    'Peso_Liquido_Total': self._format_decimal(doc.get('net_weight', 0)),
                    'Peso_Bruto_Total': self._format_decimal(doc.get('gross_weight', 0))
                })
                
                # === PAGAMENTO ===
                row.update({
                    'Forma_Pagamento': str(doc.get('payment_method', '')),
                    'Meio_Pagamento': str(doc.get('payment_type', '')),
                    'Valor_Pagamento': self._format_decimal(doc.get('payment_value', 0)),
                    'CNPJ_Credenciadora_Cartao': self._format_number_or_text(doc.get('cnpj_credenciadora', '')),
                    'Bandeira_Cartao': str(doc.get('bandeira_cartao', '')),
                    'Numero_Autorizacao_Cartao': str(doc.get('autorizacao_cartao', '')),
                    'Valor_Troco': self._format_decimal(doc.get('valor_troco', 0))
                })
                
                # === INFORMAÇÕES ADICIONAIS ===
                row.update({
                    'Informacoes_Adicionais_Interesse_Fisco': str(doc.get('tax_info', '')),
                    'Informacoes_Complementares_Contribuinte': str(doc.get('additional_info', '')),
                    'Observacoes_Gerais': str(doc.get('observations', '')),
                    'Campo_Livre_Uso_Contribuinte': str(doc.get('campo_livre', ''))
                })
                
                # === EXPORTAÇÃO ===
                row.update({
                    'Local_Embarque': str(doc.get('local_embarque', '')),
                    'Local_Despacho': str(doc.get('local_despacho', '')),
                    'UFD_Saida': str(doc.get('ufd_saida', '')),
                    'Local_Saida_Pais': str(doc.get('local_saida_pais', '')),
                    'Drawback': str(doc.get('drawback', '')),
                    'Numero_Registro_Exportacao': str(doc.get('numero_registro_exportacao', ''))
                })
                
                # === COMPRAS PÚBLICAS ===
                row.update({
                    'CNPJ_Orgao_Publico': self._format_number_or_text(doc.get('cnpj_orgao_publico', '')),
                    'Numero_Empenho': str(doc.get('numero_empenho', '')),
                    'Modalidade_Licitacao': str(doc.get('modalidade_licitacao', '')),
                    'Numero_Licitacao': str(doc.get('numero_licitacao', ''))
                })
                
                # === RESPONSÁVEL TÉCNICO ===
                row.update({
                    'CNPJ_Responsavel_Tecnico': self._format_number_or_text(doc.get('cnpj_resp_tecnico', '')),
                    'Contato_Responsavel_Tecnico': str(doc.get('contato_resp_tecnico', '')),
                    'Email_Responsavel_Tecnico': str(doc.get('email_resp_tecnico', '')),
                    'Telefone_Responsavel_Tecnico': self._format_number_or_text(doc.get('telefone_resp_tecnico', ''))
                })
                
                # === INFORMAÇÕES TÉCNICAS DO ARQUIVO ===
                row.update({
                    'Nome_Arquivo_XML': str(doc.get('file_name', '')),
                    'Tamanho_Arquivo_Bytes': self._format_decimal(doc.get('file_size', 0)),
                    'Hash_MD5_Arquivo': str(doc.get('file_hash', ''))[:32] if doc.get('file_hash') else '',
                    'Versao_Schema_XML': str(doc.get('versao_schema', '')),
                    'Algoritmo_Hash': str(doc.get('algoritmo_hash', 'MD5'))
                })
                
                # === CONTINGÊNCIA ===
                row.update({
                    'Forma_Emissao': str(doc.get('forma_emissao', '')),
                    'Justificativa_Contingencia': str(doc.get('justificativa_contingencia', '')),
                    'Data_Hora_Entrada_Contingencia': self._format_date(doc.get('data_contingencia'), date_format)
                })
                
                # === REFERENCIADOS ===
                row.update({
                    'NFe_Referenciada': str(doc.get('nfe_referenciada', '')),
                    'CNPJ_Emitente_Referenciado': self._format_number_or_text(doc.get('cnpj_emit_ref', '')),
                    'Numero_NFe_Referenciada': self._format_number_or_text(doc.get('numero_nfe_ref', '')),
                    'Serie_NFe_Referenciada': self._format_number_or_text(doc.get('serie_nfe_ref', '')),
                    'Cupom_Fiscal_Referenciado': str(doc.get('cupom_fiscal_ref', ''))
                })
                
                data.append(row)
                
            except Exception as e:
                logging.error(f"Error preparing document data for export: {e}")
                continue
        
        return data
    
    def _format_number_or_text(self, value):
        """Format value as number if it doesn't start with 0, otherwise as text"""
        if not value:
            return ''
        
        str_value = str(value).strip()
        
        # If starts with 0 and has more than 1 digit, treat as text
        if str_value.startswith('0') and len(str_value) > 1:
            return str_value
        
        # Try to format as number
        try:
            # Check if it's a valid number
            float_value = float(str_value)
            # If it's a whole number, return as integer
            if float_value.is_integer():
                return int(float_value)
            else:
                return float_value
        except (ValueError, TypeError):
            return str_value
    
    def _format_decimal(self, value):
        """Format decimal value as number for Excel"""
        try:
            if value is None or value == '':
                return 0.0
            
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _format_date(self, date_str, format_str):
        """Format date string"""
        if not date_str:
            return ''
        
        try:
            # If it's already in the right format, return as is
            if isinstance(date_str, str):
                # Try to parse ISO format first
                if 'T' in date_str or '-' in date_str:
                    from datetime import datetime
                    dt = datetime.fromisoformat(date_str.replace('T', ' ').split('.')[0])
                    return dt.strftime(format_str)
                else:
                    return date_str
            
            return str(date_str)
            
        except Exception:
            return str(date_str) if date_str else ''
    
    def _format_excel_worksheet(self, worksheet, df):
        """Apply comprehensive formatting to Excel worksheet with Brazilian fiscal standards"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
        
        # Define color scheme following Brazilian fiscal standards
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Dark blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        # Categories for different field types with specific formatting
        category_colors = {
            'identificacao': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),  # Light blue
            'datas': PatternFill(start_color="FFF2E7", end_color="FFF2E7", fill_type="solid"),  # Light orange
            'status': PatternFill(start_color="E7FFE7", end_color="E7FFE7", fill_type="solid"),  # Light green
            'emitente': PatternFill(start_color="F0E7FF", end_color="F0E7FF", fill_type="solid"),  # Light purple
            'destinatario': PatternFill(start_color="FFE7F0", end_color="FFE7F0", fill_type="solid"),  # Light pink
            'valores': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),  # Light yellow
            'tributos': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),  # Light blue 2
            'transporte': PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid"),  # Light green 2
            'pagamento': PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid"),  # Light magenta
            'tecnicas': PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray
        }
        
        # Field categorization for color coding
        field_categories = {
            'identificacao': ['ID_Interno', 'Chave_Acesso', 'Tipo_Documento', 'Modelo_Documento', 'Serie_Documento', 'Numero_Documento', 'Versao_Documento', 'Finalidade_NFe', 'Processo_Emissao', 'Versao_Processo'],
            'datas': ['Data_Emissao', 'Data_Entrada_Saida', 'Data_Processamento', 'Data_Criacao_Sistema', 'Data_Ultima_Atualizacao', 'Hora_Entrada_Saida', 'Data_Contingencia', 'Data_Autorizacao', 'Data_Hora_Entrada_Contingencia'],
            'status': ['Status_Documento', 'Situacao_Documento', 'Codigo_Status', 'Motivo_Status', 'Protocolo_Autorizacao', 'Justificativa_Cancelamento'],
            'emitente': [col for col in df.columns if 'Emitente' in col or col.startswith('CNPJ_Emitente') or col.startswith('CPF_Emitente') or col.startswith('Razao_Social_Emitente')],
            'destinatario': [col for col in df.columns if 'Destinatario' in col or col.startswith('CNPJ_Destinatario') or col.startswith('CPF_Destinatario')],
            'valores': [col for col in df.columns if col.startswith('Valor_') and not any(tax in col for tax in ['ICMS', 'IPI', 'PIS', 'COFINS', 'ISSQN', 'Retencao'])],
            'tributos': [col for col in df.columns if any(tax in col for tax in ['ICMS', 'IPI', 'PIS', 'COFINS', 'ISSQN', 'Base_Calculo', 'Aliquota', 'Retencao', 'FCP'])],
            'transporte': [col for col in df.columns if any(word in col for word in ['Transportadora', 'Frete', 'Volumes', 'Peso', 'Veiculo', 'RNTC'])],
            'pagamento': [col for col in df.columns if any(word in col for word in ['Pagamento', 'Cartao', 'Credenciadora', 'Bandeira', 'Troco'])],
            'tecnicas': [col for col in df.columns if any(word in col for word in ['Arquivo', 'Hash', 'Schema', 'Algoritmo', 'Contingencia', 'Emissao'])]
        }
        
        # Create borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Apply formatting to data cells
        for row_num in range(2, len(df) + 2):
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                # Determine category and apply color
                category_found = False
                for category, fields in field_categories.items():
                    if column in fields:
                        cell.fill = category_colors[category]
                        category_found = True
                        break
                
                if not category_found:
                    cell.fill = category_colors['tecnicas']  # Default color
                
                # Apply specific formatting based on field type
                value = cell.value
                
                # Date fields
                if 'Data_' in column or 'Hora_' in column:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if value and isinstance(value, str) and len(value) > 0:
                        cell.number_format = 'DD/MM/YYYY'
                
                # Numeric fields (values starting with Valor_, Base_Calculo_, etc.)
                elif any(prefix in column for prefix in ['Valor_', 'Base_Calculo_', 'Peso_', 'Quantidade_', 'Aliquota_']):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and value != 0:
                        if 'Aliquota_' in column:
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00'
                
                # Document numbers and codes (preserve leading zeros)
                elif any(field in column for field in ['CNPJ_', 'CPF_', 'CEP_', 'Codigo_', 'Serie_', 'Numero_', 'Inscricao_']):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '@'  # Text format to preserve leading zeros
                
                # Chave de Acesso (access key)
                elif 'Chave_Acesso' in column:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.number_format = '@'
                    cell.font = Font(name="Courier New", size=9)  # Monospace for better readability
                
                # Text fields
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths with intelligent sizing
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate optimal width
            max_length = len(str(column))  # Header length
            
            # Sample some data to determine width
            for row_num in range(2, min(len(df) + 2, 52)):  # Check up to 50 rows
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    max_length = max(max_length, cell_length)
            
            # Set intelligent width limits based on field type
            if 'Chave_Acesso' in column:
                width = 50  # Access key needs full width
            elif any(field in column for field in ['CNPJ_', 'CPF_']):
                width = min(max_length + 2, 20)
            elif any(field in column for field in ['CEP_', 'Telefone_']):
                width = min(max_length + 2, 15)
            elif 'Email_' in column:
                width = min(max_length + 2, 30)
            elif any(field in column for field in ['Endereco_', 'Razao_Social_', 'Nome_Fantasia_']):
                width = min(max_length + 2, 35)
            elif any(field in column for field in ['Valor_', 'Base_Calculo_']):
                width = min(max_length + 2, 18)
            elif 'Data_' in column:
                width = 12
            else:
                width = min(max_length + 2, 25)
            
            # Apply minimum and maximum width constraints
            width = max(8, min(width, 60))
            worksheet.column_dimensions[column_letter].width = width
        
        # Freeze panes (first row and first 3 columns for navigation)
        worksheet.freeze_panes = "D2"
        
        # Add auto-filter to enable filtering
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # Set row height for better readability
        worksheet.row_dimensions[1].height = 30  # Header row
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20  # Data rows
    
    def _add_summary_sheet(self, writer, df):
        """Add summary sheet to Excel file"""
        summary_data = {
            'Estatística': [
                'Total de Documentos',
                'Valor Total',
                'Média por Documento',
                'Total de Impostos',
                'Documentos por Tipo'
            ],
            'Valor': [
                len(df),
                f"{df['Valor Total'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if 'Valor Total' in df.columns else '0,00',
                f"{df['Valor Total'].mean():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if 'Valor Total' in df.columns else '0,00',
                f"{df['Total de Impostos'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if 'Total de Impostos' in df.columns else '0,00',
                ''
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumo', index=False)

    def _prepare_products_data(self):
        """Prepare products/items data for export with comprehensive information"""
        data = []
        date_format = self.export_config.get('date_format', '%d/%m/%Y')
        
        for doc in self.documents:
            try:
                # Get document items
                items = doc.get('items', [])
                
                # Extract XML number from filename
                file_name = str(doc.get('file_name', ''))
                xml_number = ''
                if file_name:
                    # Try to extract the access key from file name (44-digit number)
                    import re
                    access_key_match = re.search(r'\d{44}', file_name)
                    if access_key_match:
                        xml_number = access_key_match.group()
                    else:
                        xml_number = file_name.replace('.xml', '').replace('-procNFe', '')
                
                # If no items, create a single row for the document
                if not items:
                    items = [{}]
                
                for item in items:
                    row = {}
                    
                    # === IDENTIFICAÇÃO DO PRODUTO/ITEM ===
                    row.update({
                        'Numero_XML': xml_number,
                        'Nome_Arquivo_XML': file_name,
                        'Codigo_Item': str(item.get('item_code', '')),
                        'Codigo_EAN_GTIN': str(item.get('ean_gtin', '')),
                        'Descricao_Produto': str(item.get('item_description', '')),
                        'NCM': str(item.get('ncm_code', '')),
                        'CEST': str(item.get('cest_code', '')),
                        'CFOP': str(item.get('cfop', '')),
                        'Unidade_Comercial': str(item.get('commercial_unit', '')),
                        'Ex_TIPI': str(item.get('ex_tipi', '')),
                        'Genero': str(item.get('genero', '')),
                        'Especificacao': str(item.get('especificacao', ''))
                    })
                    
                    # === QUANTIDADES E VALORES ===
                    row.update({
                        'Quantidade_Comercial': self._format_decimal(item.get('quantity', 0)),
                        'Valor_Unitario_Comercial': self._format_decimal(item.get('unit_value', 0)),
                        'Valor_Total_Produto': self._format_decimal(item.get('total_value', 0)),
                        'Valor_Frete': self._format_decimal(item.get('freight_value', 0)),
                        'Valor_Seguro': self._format_decimal(item.get('insurance_value', 0)),
                        'Valor_Desconto': self._format_decimal(item.get('discount_value', 0)),
                        'Outras_Despesas': self._format_decimal(item.get('other_expenses', 0)),
                        'Valor_Aproximado_Tributos': self._format_decimal(item.get('valor_aproximado_tributos', 0))
                    })
                    
                    # === INFORMAÇÕES TRIBUTÁRIAS - ICMS ===
                    row.update({
                        'ICMS_Origem': str(item.get('icms_origin', '')),
                        'ICMS_CST': str(item.get('icms_cst', '')),
                        'ICMS_Modalidade_BC': str(item.get('icms_modbc', '')),
                        'ICMS_Base_Calculo': self._format_decimal(item.get('icms_bc', 0)),
                        'ICMS_Aliquota': self._format_decimal(item.get('icms_rate', 0)),
                        'ICMS_Valor': self._format_decimal(item.get('icms_value', 0)),
                        'ICMS_ST_Modalidade_BC': str(item.get('icms_modbc_st', '')),
                        'ICMS_ST_Base_Calculo': self._format_decimal(item.get('icms_bc_st', 0)),
                        'ICMS_ST_Aliquota': self._format_decimal(item.get('icms_rate_st', 0)),
                        'ICMS_ST_Valor': self._format_decimal(item.get('icms_value_st', 0)),
                        'ICMS_ST_BC_Retencao': self._format_decimal(item.get('icms_bc_st_ret', 0)),
                        'ICMS_ST_Valor_Retencao': self._format_decimal(item.get('icms_value_st_ret', 0)),
                        'ICMS_FCP_Base_Calculo': self._format_decimal(item.get('icms_bc_fcp', 0)),
                        'ICMS_FCP_Aliquota': self._format_decimal(item.get('icms_rate_fcp', 0)),
                        'ICMS_FCP_Valor': self._format_decimal(item.get('icms_value_fcp', 0))
                    })
                    
                    # === INFORMAÇÕES TRIBUTÁRIAS - IPI ===
                    row.update({
                        'IPI_Classe_Enquadramento': str(item.get('ipi_cl_enq', '')),
                        'IPI_CNPJ_Produtor': self._format_cnpj_as_text(item.get('ipi_cnpj_prod', '')),
                        'IPI_Codigo_Selo': str(item.get('ipi_c_selo', '')),
                        'IPI_Quantidade_Selo': self._format_decimal(item.get('ipi_q_selo', 0)),
                        'IPI_Codigo_Enquadramento': str(item.get('ipi_c_enq', '')),
                        'IPI_CST': str(item.get('ipi_cst', '')),
                        'IPI_Base_Calculo': self._format_decimal(item.get('ipi_bc', 0)),
                        'IPI_Aliquota': self._format_decimal(item.get('ipi_rate', 0)),
                        'IPI_Valor': self._format_decimal(item.get('ipi_value', 0)),
                        'IPI_Base_Calculo_Unitaria': self._format_decimal(item.get('ipi_bc_unit', 0)),
                        'IPI_Valor_Unitario': self._format_decimal(item.get('ipi_unit_value', 0)),
                        'IPI_Quantidade_Unidade': self._format_decimal(item.get('ipi_unit_qty', 0))
                    })
                    
                    # === INFORMAÇÕES TRIBUTÁRIAS - PIS ===
                    row.update({
                        'PIS_CST': str(item.get('pis_cst', '')),
                        'PIS_Base_Calculo': self._format_decimal(item.get('pis_bc', 0)),
                        'PIS_Aliquota': self._format_decimal(item.get('pis_rate', 0)),
                        'PIS_Valor': self._format_decimal(item.get('pis_value', 0)),
                        'PIS_Quantidade_Vendida': self._format_decimal(item.get('pis_qty_sold', 0)),
                        'PIS_Aliquota_Unitaria': self._format_decimal(item.get('pis_aliq_unit', 0)),
                        'PIS_Valor_Unitario': self._format_decimal(item.get('pis_value_unit', 0)),
                        'PIS_ST_Base_Calculo': self._format_decimal(item.get('pis_st_bc', 0)),
                        'PIS_ST_Aliquota': self._format_decimal(item.get('pis_st_rate', 0)),
                        'PIS_ST_Valor': self._format_decimal(item.get('pis_st_value', 0))
                    })
                    
                    # === INFORMAÇÕES TRIBUTÁRIAS - COFINS ===
                    row.update({
                        'COFINS_CST': str(item.get('cofins_cst', '')),
                        'COFINS_Base_Calculo': self._format_decimal(item.get('cofins_bc', 0)),
                        'COFINS_Aliquota': self._format_decimal(item.get('cofins_rate', 0)),
                        'COFINS_Valor': self._format_decimal(item.get('cofins_value', 0)),
                        'COFINS_Quantidade_Vendida': self._format_decimal(item.get('cofins_qty_sold', 0)),
                        'COFINS_Aliquota_Unitaria': self._format_decimal(item.get('cofins_aliq_unit', 0)),
                        'COFINS_Valor_Unitario': self._format_decimal(item.get('cofins_value_unit', 0)),
                        'COFINS_ST_Base_Calculo': self._format_decimal(item.get('cofins_st_bc', 0)),
                        'COFINS_ST_Aliquota': self._format_decimal(item.get('cofins_st_rate', 0)),
                        'COFINS_ST_Valor': self._format_decimal(item.get('cofins_st_value', 0))
                    })
                    
                    # === OUTROS IMPOSTOS ===
                    row.update({
                        'II_Base_Calculo': self._format_decimal(item.get('ii_bc', 0)),
                        'II_Despesas_Aduaneiras': self._format_decimal(item.get('ii_despesas_aduaneiras', 0)),
                        'II_Valor': self._format_decimal(item.get('ii_value', 0)),
                        'II_IOF': self._format_decimal(item.get('ii_iof', 0))
                    })
                    
                    # === INFORMAÇÕES DO DOCUMENTO ===
                    row.update({
                        'ID_Documento': str(doc.get('id', '')),
                        'Tipo_Documento': str(doc.get('document_type', '')).upper(),
                        'Numero_Documento': self._format_number_or_text(doc.get('document_number', '')),
                        'Serie_Documento': self._format_number_or_text(doc.get('series', '')),
                        'Data_Emissao': self._format_date(doc.get('issue_date'), date_format),
                        'CNPJ_Emitente': self._format_cnpj_as_text(doc.get('cnpj_issuer', '')),
                        'Razao_Social_Emitente': str(doc.get('issuer_name', '')),
                        'CNPJ_Destinatario': self._format_cnpj_as_text(doc.get('cnpj_recipient', '')),
                        'Razao_Social_Destinatario': str(doc.get('recipient_name', '')),
                        'Valor_Total_Documento': self._format_decimal(doc.get('total_value', 0)),
                        'Status_Documento': str(doc.get('status', '')).title()
                    })
                    
                    # === INFORMAÇÕES ADICIONAIS DO ITEM ===
                    row.update({
                        'Informacoes_Adicionais_Item': str(item.get('additional_info', '')),
                        'Observacoes_Item': str(item.get('observacoes', '')),
                        'Classificacao_Fiscal': str(item.get('tax_classification', ''))
                    })
                    
                    data.append(row)
                    
            except Exception as e:
                logging.error(f"Error preparing product data for export: {e}")
                continue
        
        return data
    
    def _format_cnpj_as_text(self, value):
        """Format CNPJ/CPF specifically as text to preserve leading zeros"""
        if not value:
            return ''
        
        # Always return as string to preserve leading zeros
        return str(value).strip()

    def _export_products_excel(self):
        """Export products/items to Excel format with comprehensive formatting"""
        if not EXCEL_AVAILABLE:
            # Try alternative export method for products
            return self._export_products_excel_alternative()
        
        try:
            self.progress_updated.emit(10, "Preparando dados dos produtos para Excel...")
            
            # Prepare products data
            data = self._prepare_products_data()
            if not data:
                return False, "Nenhum produto encontrado para exportar"
            
            df = pd.DataFrame(data)
            
            self.progress_updated.emit(30, "Criando arquivo Excel dos produtos...")
            
            # Create workbook with error handling
            try:
                with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                    # Write products data
                    df.to_excel(writer, sheet_name='Produtos', index=False)
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Produtos']
                    
                    self.progress_updated.emit(60, "Formatando planilha de produtos...")
                    
                    # Apply specific formatting for products
                    self._format_products_excel_worksheet(worksheet, df)
                    
                    self.progress_updated.emit(90, "Finalizando arquivo...")
                    
            except Exception as e:
                # If openpyxl fails, try xlsxwriter
                try:
                    with pd.ExcelWriter(self.output_path, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='Produtos', index=False)
                        self.progress_updated.emit(90, "Finalizando arquivo...")
                except Exception as e2:
                    return False, f"Erro ao criar arquivo Excel de produtos: {str(e2)}"
            
            self.progress_updated.emit(100, "Exportação de produtos concluída!")
            return True, f"Arquivo Excel de produtos criado: {self.output_path}"
            
        except Exception as e:
            logging.error(f"Products Excel export error: {e}")
            return False, f"Erro ao exportar produtos para Excel: {str(e)}"

    def _export_products_excel_alternative(self):
        """Alternative products export using CSV format when Excel libraries are not available"""
        try:
            self.progress_updated.emit(10, "Bibliotecas Excel não disponíveis, usando formato CSV para produtos...")
            
            # Change output path to CSV
            csv_path = str(self.output_path).replace('.xlsx', '_produtos.csv').replace('.xls', '_produtos.csv')
            
            data = self._prepare_products_data()
            if not data:
                return False, "Nenhum produto encontrado para exportar"
            
            self.progress_updated.emit(50, "Escrevendo arquivo CSV de produtos...")
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
                    
                    writer.writeheader()
                    
                    for i, row in enumerate(data):
                        writer.writerow(row)
                        if i % 100 == 0:
                            progress = 50 + int((i / len(data)) * 40)
                            self.progress_updated.emit(progress, f"Processando produto {i+1}...")
            
            self.progress_updated.emit(100, "Exportação de produtos concluída!")
            return True, f"Arquivo CSV de produtos criado (Excel não disponível): {csv_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar produtos para CSV alternativo: {str(e)}"
    
    def _format_products_excel_worksheet(self, worksheet, df):
        """Apply specific formatting to products Excel worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define color scheme for products
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Dark blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        # Categories for different field types with specific formatting
        category_colors = {
            'identificacao': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),  # Light blue
            'valores': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),  # Light yellow
            'tributos_icms': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),  # Light blue 2
            'tributos_ipi': PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid"),  # Light green
            'tributos_pis': PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid"),  # Light magenta
            'tributos_cofins': PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),  # Light orange
            'documento': PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),  # Light gray
            'adicional': PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")  # Light gray 2
        }
        
        # Field categorization for color coding
        field_categories = {
            'identificacao': ['Numero_XML', 'Nome_Arquivo_XML', 'Codigo_Item', 'Codigo_EAN_GTIN', 'Descricao_Produto', 'NCM', 'CEST', 'CFOP', 'Unidade_Comercial', 'Ex_TIPI', 'Genero', 'Especificacao'],
            'valores': ['Quantidade_Comercial', 'Valor_Unitario_Comercial', 'Valor_Total_Produto', 'Valor_Frete', 'Valor_Seguro', 'Valor_Desconto', 'Outras_Despesas', 'Valor_Aproximado_Tributos'],
            'tributos_icms': [col for col in df.columns if col.startswith('ICMS_')],
            'tributos_ipi': [col for col in df.columns if col.startswith('IPI_')],
            'tributos_pis': [col for col in df.columns if col.startswith('PIS_')],
            'tributos_cofins': [col for col in df.columns if col.startswith('COFINS_')],
            'documento': ['ID_Documento', 'Tipo_Documento', 'Numero_Documento', 'Serie_Documento', 'Data_Emissao', 'CNPJ_Emitente', 'Razao_Social_Emitente', 'CNPJ_Destinatario', 'Razao_Social_Destinatario', 'Valor_Total_Documento', 'Status_Documento'],
            'adicional': ['Informacoes_Adicionais_Item', 'Observacoes_Item', 'Classificacao_Fiscal']
        }
        
        # Create borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Apply formatting to data cells
        for row_num in range(2, len(df) + 2):
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                # Determine category and apply color
                category_found = False
                for category, fields in field_categories.items():
                    if column in fields:
                        cell.fill = category_colors[category]
                        category_found = True
                        break
                
                if not category_found:
                    cell.fill = category_colors['adicional']  # Default color
                
                # Apply specific formatting based on field type
                value = cell.value
                
                # Numeric fields (values and calculations)
                if any(prefix in column for prefix in ['Valor_', 'Quantidade_', 'Aliquota', 'Base_Calculo']):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and value != 0:
                        if 'Aliquota' in column or 'Percentual' in column:
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00'
                
                # CNPJ fields (preserve leading zeros)
                elif 'CNPJ_' in column or 'CPF_' in column:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '@'  # Text format to preserve leading zeros
                
                # Document numbers and codes
                elif any(field in column for field in ['Codigo_', 'Numero_', 'Serie_', 'NCM', 'CEST', 'CFOP', 'CST']):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '@'  # Text format to preserve leading zeros
                
                # XML Number/Access Key
                elif 'Numero_XML' in column:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.number_format = '@'
                    cell.font = Font(name="Courier New", size=9)  # Monospace for better readability
                
                # Date fields
                elif 'Data_' in column:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = 'DD/MM/YYYY'
                
                # Text fields
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths with intelligent sizing
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate optimal width
            max_length = len(str(column))  # Header length
            
            # Sample some data to determine width
            for row_num in range(2, min(len(df) + 2, 52)):  # Check up to 50 rows
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    max_length = max(max_length, cell_length)
            
            # Set intelligent width limits based on field type
            if 'Numero_XML' in column:
                width = 50  # XML number needs full width
            elif 'Descricao_Produto' in column:
                width = min(max_length + 2, 50)
            elif any(field in column for field in ['CNPJ_', 'CPF_']):
                width = min(max_length + 2, 20)
            elif any(field in column for field in ['Razao_Social_', 'Nome_']):
                width = min(max_length + 2, 35)
            elif any(field in column for field in ['Valor_', 'Base_Calculo_', 'Quantidade_']):
                width = min(max_length + 2, 18)
            elif 'Data_' in column:
                width = 12
            else:
                width = min(max_length + 2, 25)
            
            # Apply minimum and maximum width constraints
            width = max(8, min(width, 60))
            worksheet.column_dimensions[column_letter].width = width
        
        # Freeze panes (first row and first 3 columns for navigation)
        worksheet.freeze_panes = "D2"
        
        # Add auto-filter to enable filtering
        worksheet.auto_filter.ref = worksheet.dimensions

    def _export_products_csv(self):
        """Export products/items to CSV format"""
        try:
            self.progress_updated.emit(20, "Preparando dados dos produtos para CSV...")
            
            data = self._prepare_products_data()
            if not data:
                return False, "Nenhum produto encontrado para exportar"
            
            self.progress_updated.emit(50, "Escrevendo arquivo CSV de produtos...")
            
            with open(self.output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, 
                                          delimiter=self.export_config.get('csv_delimiter', ';'))
                    
                    if self.export_config.get('include_header', True):
                        writer.writeheader()
                    
                    for i, row in enumerate(data):
                        writer.writerow(row)
                        if i % 100 == 0:
                            progress = 50 + int((i / len(data)) * 40)
                            self.progress_updated.emit(progress, f"Processando produto {i+1}...")
            
            self.progress_updated.emit(100, "Exportação de produtos concluída!")
            return True, f"Arquivo CSV de produtos criado: {self.output_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar produtos para CSV: {str(e)}"
    
    def _export_products_json(self):
        """Export products/items to JSON format"""
        try:
            self.progress_updated.emit(20, "Preparando dados dos produtos para JSON...")
            
            data = self._prepare_products_data()
            if not data:
                return False, "Nenhum produto encontrado para exportar"
            
            self.progress_updated.emit(50, "Escrevendo arquivo JSON de produtos...")
            
            # Prepare JSON structure
            export_data = {
                'export_info': {
                    'timestamp': datetime.now().isoformat(),
                    'total_products': len(data),
                    'export_type': 'products',
                    'format_version': '1.0'
                },
                'products': data
            }
            
            with open(self.output_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)
            
            self.progress_updated.emit(100, "Exportação de produtos concluída!")
            return True, f"Arquivo JSON de produtos criado: {self.output_path}"
            
        except Exception as e:
            return False, f"Erro ao exportar produtos para JSON: {str(e)}"


class ExportDialog(QDialog):
    """Advanced export dialog with support for documents and products"""
    
    def __init__(self, documents, config, parent=None):
        super().__init__(parent)
        
        self.documents = documents
        self.config = config
        self.export_worker = None
        
        self._setup_dialog()
        self._create_ui()
        self._load_settings()
    
    def _setup_dialog(self):
        """Setup dialog properties"""
        self.setWindowTitle("Exportar Dados")
        self.setModal(True)
        self.resize(650, 750)
        
        # Main layout
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(15, 15, 15, 15)
    
    def _create_ui(self):
        """Create user interface"""
        # Document info
        info_group = QGroupBox("Informações da Exportação")
        info_layout = QFormLayout(info_group)
        
        self.doc_count_label = QLabel(f"{len(self.documents)} documentos selecionados")
        self.doc_count_label.setFont(QFont("", 10, QFont.Bold))
        info_layout.addRow("Total:", self.doc_count_label)
        
        # Calculate totals
        total_value = sum(doc.get('total_value', 0) for doc in self.documents)
        total_tax = sum(doc.get('tax_value', 0) for doc in self.documents)
        
        # Format without R$ symbol
        total_value_formatted = f"{total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        total_tax_formatted = f"{total_tax:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        self.total_value_label = QLabel(total_value_formatted)
        info_layout.addRow("Valor Total:", self.total_value_label)
        
        self.total_tax_label = QLabel(total_tax_formatted)
        info_layout.addRow("Total de Impostos:", self.total_tax_label)
        
        self.main_layout.addWidget(info_group)
        
        # Export type selection
        type_group = QGroupBox("Tipo de Exportação")
        type_layout = QVBoxLayout(type_group)
        
        self.export_type_combo = QComboBox()
        self.export_type_combo.addItems(["Documentos (resumo)", "Produtos/Itens (detalhado)"])
        self.export_type_combo.currentTextChanged.connect(self._on_export_type_changed)
        type_layout.addWidget(QLabel("Tipo de dados:"))
        type_layout.addWidget(self.export_type_combo)
        
        self.main_layout.addWidget(type_group)
        
        # Create tabs
        self.tab_widget = QTabWidget()
        self.main_layout.addWidget(self.tab_widget)
        
        self._create_format_tab()
        self._create_fields_tab()
        self._create_options_tab()
        
        # Progress bar (hidden initially)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.main_layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel()
        self.status_label.setVisible(False)
        self.main_layout.addWidget(self.status_label)
        
        # Buttons
        self._create_buttons()
    
    def _create_format_tab(self):
        """Create format selection tab"""
        format_widget = QFrame()
        layout = QVBoxLayout(format_widget)
        
        # Output format
        format_group = QGroupBox("Formato de Saída")
        format_layout = QVBoxLayout(format_group)
        
        self.format_combo = QComboBox()
        formats = ["Excel (.xlsx)", "CSV", "JSON"]
        if not EXCEL_AVAILABLE:
            formats = ["CSV", "JSON"]
        
        self.format_combo.addItems(formats)
        self.format_combo.currentTextChanged.connect(self._on_format_changed)
        format_layout.addWidget(QLabel("Formato:"))
        format_layout.addWidget(self.format_combo)
        
        layout.addWidget(format_group)
        
        # Output file
        file_group = QGroupBox("Arquivo de Saída")
        file_layout = QVBoxLayout(file_group)
        
        file_path_layout = QHBoxLayout()
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Selecione o local para salvar...")
        browse_btn = QPushButton("Procurar...")
        browse_btn.clicked.connect(self._browse_output_file)
        
        file_path_layout.addWidget(self.file_path_edit)
        file_path_layout.addWidget(browse_btn)
        file_layout.addLayout(file_path_layout)
        
        layout.addWidget(file_group)
        
        # Format-specific options
        self.format_options_group = QGroupBox("Opções do Formato")
        self.format_options_layout = QVBoxLayout(self.format_options_group)
        layout.addWidget(self.format_options_group)
        
        self._update_format_options()
        
        layout.addStretch()
        
        self.tab_widget.addTab(format_widget, "Formato")
    
    def _create_fields_tab(self):
        """Create fields selection tab"""
        fields_widget = QFrame()
        layout = QVBoxLayout(fields_widget)
        
        # Field selection
        fields_group = QGroupBox("Campos a Incluir")
        fields_layout = QVBoxLayout(fields_group)
        
        # Select all/none buttons
        select_buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("Selecionar Todos")
        select_all_btn.clicked.connect(self._select_all_fields)
        select_none_btn = QPushButton("Desmarcar Todos")
        select_none_btn.clicked.connect(self._select_no_fields)
        
        select_buttons_layout.addWidget(select_all_btn)
        select_buttons_layout.addWidget(select_none_btn)
        select_buttons_layout.addStretch()
        fields_layout.addLayout(select_buttons_layout)
        
        # Field checkboxes
        self.field_checks = {}
        
        field_groups = {
            'Informações Básicas': ['basic'],
            'Dados do Emitente': ['issuer'],
            'Dados do Destinatário': ['recipient'],
            'Informações Financeiras': ['financial'],
            'Informações Fiscais': ['taxes'],
            'Informações Técnicas': ['technical']
        }
        
        for group_name, fields in field_groups.items():
            group_check = QCheckBox(group_name)
            group_check.setChecked(True)
            group_check.setFont(QFont("", 9, QFont.Bold))
            
            for field in fields:
                self.field_checks[field] = group_check
            
            fields_layout.addWidget(group_check)
        
        layout.addWidget(fields_group)
        layout.addStretch()
        
        self.tab_widget.addTab(fields_widget, "Campos")
    
    def _create_options_tab(self):
        """Create export options tab"""
        options_widget = QFrame()
        layout = QVBoxLayout(options_widget)
        
        # Date and number formatting
        format_group = QGroupBox("Formatação")
        format_layout = QFormLayout(format_group)
        
        self.date_format_combo = QComboBox()
        self.date_format_combo.addItems(["DD/MM/AAAA", "AAAA-MM-DD", "MM/DD/AAAA"])
        format_layout.addRow("Formato de data:", self.date_format_combo)
        
        self.decimal_separator_combo = QComboBox()
        self.decimal_separator_combo.addItems([",", "."])
        format_layout.addRow("Separador decimal:", self.decimal_separator_combo)
        
        layout.addWidget(format_group)
        
        # Additional options
        additional_group = QGroupBox("Opções Adicionais")
        additional_layout = QVBoxLayout(additional_group)
        
        self.include_header_check = QCheckBox("Incluir cabeçalho")
        self.include_header_check.setChecked(True)
        additional_layout.addWidget(self.include_header_check)
        
        self.include_summary_check = QCheckBox("Incluir resumo (apenas Excel)")
        self.include_summary_check.setChecked(True)
        additional_layout.addWidget(self.include_summary_check)
        
        self.open_after_export_check = QCheckBox("Abrir arquivo após exportação")
        self.open_after_export_check.setChecked(True)
        additional_layout.addWidget(self.open_after_export_check)
        
        layout.addWidget(additional_group)
        
        layout.addStretch()
        
        self.tab_widget.addTab(options_widget, "Opções")
    
    def _create_buttons(self):
        """Create dialog buttons"""
        buttons_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("Exportar")
        self.export_btn.setDefault(True)
        self.export_btn.clicked.connect(self._start_export)
        
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.clicked.connect(self.reject)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(self.export_btn)
        
        self.main_layout.addLayout(buttons_layout)
    
    def _load_settings(self):
        """Load export settings from config"""
        export_settings = self.config.get_export_config()
        
        # Set default format
        default_format = export_settings.get('default_format', 'Excel (.xlsx)')
        index = self.format_combo.findText(default_format)
        if index >= 0:
            self.format_combo.setCurrentIndex(index)
        
        # Set default path
        default_path = export_settings.get('default_path', '')
        if default_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}"
            self.file_path_edit.setText(str(Path(default_path) / filename))
        
        # Load other settings
        self.date_format_combo.setCurrentText(export_settings.get('date_format', 'DD/MM/AAAA'))
        self.decimal_separator_combo.setCurrentText(export_settings.get('decimal_separator', ','))
        self.include_header_check.setChecked(export_settings.get('include_header', True))
    
    def _on_export_type_changed(self):
        """Handle export type change"""
        export_type = self.export_type_combo.currentText()
        if "Produtos" in export_type:
            self.setWindowTitle("Exportar Produtos/Itens")
        else:
            self.setWindowTitle("Exportar Documentos")
    
    def _on_format_changed(self):
        """Handle format change"""
        self._update_format_options()
        self._update_file_extension()
    
    def _update_format_options(self):
        """Update format-specific options"""
        # Clear existing options
        for i in reversed(range(self.format_options_layout.count())):
            child = self.format_options_layout.itemAt(i).widget()
            if child:
                child.setParent(None)
        
        current_format = self.format_combo.currentText().lower()
        
        if 'excel' in current_format:
            # Excel options
            self.excel_autofit_check = QCheckBox("Ajustar largura das colunas automaticamente")
            self.excel_autofit_check.setChecked(True)
            self.format_options_layout.addWidget(self.excel_autofit_check)
            
            self.excel_freeze_header_check = QCheckBox("Congelar linha de cabeçalho")
            self.excel_freeze_header_check.setChecked(True)
            self.format_options_layout.addWidget(self.excel_freeze_header_check)
            
        elif 'csv' in current_format:
            # CSV options
            delimiter_layout = QHBoxLayout()
            delimiter_layout.addWidget(QLabel("Delimitador:"))
            self.csv_delimiter_combo = QComboBox()
            self.csv_delimiter_combo.addItems([",", ";", "\t"])
            self.csv_delimiter_combo.setCurrentText(";")  # Default to semicolon for Brazilian format
            delimiter_layout.addWidget(self.csv_delimiter_combo)
            delimiter_layout.addStretch()
            
            self.format_options_layout.addLayout(delimiter_layout)
    
    def _update_file_extension(self):
        """Update file extension based on format"""
        current_path = self.file_path_edit.text()
        if current_path:
            path = Path(current_path)
            stem = path.stem
            
            current_format = self.format_combo.currentText().lower()
            
            if 'excel' in current_format:
                extension = '.xlsx'
            elif 'csv' in current_format:
                extension = '.csv'
            elif 'json' in current_format:
                extension = '.json'
            else:
                extension = '.txt'
            
            new_path = path.parent / (stem + extension)
            self.file_path_edit.setText(str(new_path))
    
    def _browse_output_file(self):
        """Browse for output file"""
        current_format = self.format_combo.currentText()
        export_type = self.export_type_combo.currentText()
        
        if 'Excel' in current_format:
            filter_str = "Excel Files (*.xlsx);;All Files (*)"
            default_ext = ".xlsx"
        elif 'CSV' in current_format:
            filter_str = "CSV Files (*.csv);;All Files (*)"
            default_ext = ".csv"
        elif 'JSON' in current_format:
            filter_str = "JSON Files (*.json);;All Files (*)"
            default_ext = ".json"
        else:
            filter_str = "All Files (*)"
            default_ext = ".txt"
        
        current_path = self.file_path_edit.text()
        if not current_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = "produtos" if "Produtos" in export_type else "documentos"
            current_path = f"{prefix}_{timestamp}{default_ext}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Arquivo de Exportação",
            current_path, filter_str
        )
        
        if file_path:
            self.file_path_edit.setText(file_path)
    
    def _select_all_fields(self):
        """Select all field checkboxes"""
        for check in self.field_checks.values():
            check.setChecked(True)
    
    def _select_no_fields(self):
        """Deselect all field checkboxes"""
        for check in self.field_checks.values():
            check.setChecked(False)
    
    def _start_export(self):
        """Start export process"""
        # Validate inputs
        if not self.file_path_edit.text().strip():
            QMessageBox.warning(self, "Aviso", "Selecione um arquivo de saída!")
            return
        
        selected_fields = [field for field, check in self.field_checks.items() 
                          if check.isChecked()]
        
        if not selected_fields:
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um grupo de campos!")
            return
        
        # Determine export type
        export_type_text = self.export_type_combo.currentText()
        export_type = 'products' if 'Produtos' in export_type_text else 'documents'
        
        # Prepare export configuration
        export_config = {
            'format': self._get_format_key(),
            'export_type': export_type,
            'include_fields': selected_fields,
            'date_format': self._get_date_format(),
            'decimal_separator': self.decimal_separator_combo.currentText(),
            'include_header': self.include_header_check.isChecked(),
            'include_summary': self.include_summary_check.isChecked()
        }
        
        # Add format-specific options
        current_format = self.format_combo.currentText().lower()
        
        if 'excel' in current_format:
            export_config.update({
                'excel_autofit_columns': getattr(self, 'excel_autofit_check', None) and 
                                       self.excel_autofit_check.isChecked(),
                'excel_freeze_header': getattr(self, 'excel_freeze_header_check', None) and 
                                     self.excel_freeze_header_check.isChecked()
            })
        elif 'csv' in current_format:
            export_config['csv_delimiter'] = getattr(self, 'csv_delimiter_combo', None) and \
                                           self.csv_delimiter_combo.currentText() or ';'
        
        # Start export
        output_path = self.file_path_edit.text()
        
        # Show progress
        self.progress_bar.setVisible(True)
        self.status_label.setVisible(True)
        self.export_btn.setEnabled(False)
        
        # Start worker thread
        self.export_worker = ExportWorker(self.documents, export_config, output_path)
        self.export_worker.progress_updated.connect(self._update_progress)
        self.export_worker.finished.connect(self._export_finished)
        self.export_worker.error_occurred.connect(self._export_error)
        self.export_worker.start()
    
    def _get_format_key(self):
        """Get format key from combo text"""
        text = self.format_combo.currentText().lower()
        if 'excel' in text:
            return 'excel'
        elif 'csv' in text:
            return 'csv'
        elif 'json' in text:
            return 'json'
        return 'csv'
    
    def _get_date_format(self):
        """Get date format string"""
        text = self.date_format_combo.currentText()
        if text == "DD/MM/AAAA":
            return "%d/%m/%Y"
        elif text == "AAAA-MM-DD":
            return "%Y-%m-%d"
        elif text == "MM/DD/AAAA":
            return "%m/%d/%Y"
        return "%d/%m/%Y"
    
    def _update_progress(self, progress, message):
        """Update export progress"""
        self.progress_bar.setValue(progress)
        self.status_label.setText(message)
    
    def _export_finished(self, success, message):
        """Handle export completion"""
        self.progress_bar.setVisible(False)
        self.status_label.setVisible(False)
        self.export_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "Sucesso", message)
            
            # Open file if requested
            if self.open_after_export_check.isChecked():
                try:
                    import os
                    os.startfile(self.file_path_edit.text())
                except:
                    pass  # Ignore if can't open
            
            self.accept()
        else:
            QMessageBox.critical(self, "Erro", message)
    
    def _export_error(self, error):
        """Handle export error"""
        self.progress_bar.setVisible(False)
        self.status_label.setVisible(False)
        self.export_btn.setEnabled(True)
        
        QMessageBox.critical(self, "Erro", f"Erro durante exportação:\n{error}")